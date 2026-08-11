"""Generate Excel template workbooks with structured, multi-row headers
derived from a ``TableSchema``.

Header layout (total rows = ``schema.header_rows`` = max_nesting_depth + 1):

1. The first ``max_nesting_depth`` rows are **name+type rows**. Each cell uses
   openpyxl's ``CellRichText`` to stack two text runs:
   - line 1: field name (12pt bold white, on the cell's name fill)
   - line 2: type annotation (9pt italic light green ``D8F3DC``)

   Non-struct fields write into the first name row and merge vertically
   downward through the remaining name rows. Struct fields merge horizontally
   across the span of their sub-fields and recurse one row deeper. Struct
   cells display ``field.name`` as their type, matching the FBS-generated
   table name (schema names pass through verbatim).
2. The **comment row** (last) shows ``field.comment``.
"""

from __future__ import annotations

import logging
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime, timezone
from importlib import metadata as importlib_metadata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.packaging.custom import (
    DateTimeProperty,
    IntProperty,
    StringProperty,
)
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ct.schema.hashing import compute_schema_hash
from ct.schema.models import FieldDef, TableSchema
from ct.schema.type_traits import TYPE_TRAITS

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

# Cell-level font (used by openpyxl for default rendering; rich-text runs
# carry their own InlineFont below).
_HEADER_FONT = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
_COMMENT_FONT = Font(name="微软雅黑", italic=True, size=9, color="888888")

# Rich-text inline fonts for the two stacked runs inside each header cell.
_NAME_RUN_FONT = InlineFont(rFont="Segoe UI", b=True, sz=12, color="FFFFFF")
_TYPE_RUN_FONT = InlineFont(rFont="Consolas", i=True, sz=9, color="D8F3DC")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Header row fills — deep green hierarchy (dark → mid → primary)
_NORMAL_FILL  = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")  # deep forest green
_STRUCT_FILL  = PatternFill(start_color="40916C", end_color="40916C", fill_type="solid")  # mid green
_PRIMARY_FILL = PatternFill(start_color="C9A227", end_color="C9A227", fill_type="solid")  # warm gold (primary key)
_COMMENT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # light gray

# Data-row zebra fills
_ZEBRA_FILL = PatternFill(start_color="EDF7EE", end_color="EDF7EE", fill_type="solid")  # pale green (even rows)
_WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # white (odd rows)
_COL_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)

# Explicit row height for name+type rows so the two-line rich text is not
# clipped on viewers that ignore wrap_text auto-fit (WPS, Excel for Mac).
_NAME_ROW_HEIGHT = 36


# ---------------------------------------------------------------------------
# Collect all leaf fields with their 1-based column positions
# ---------------------------------------------------------------------------

def _collect_leaf_fields(fields: list[FieldDef], start_col: int) -> list[tuple[FieldDef, int]]:
    """Return (leaf_field, col_index) pairs for all non-struct leaf fields."""
    result: list[tuple[FieldDef, int]] = []
    col = start_col
    for field in fields:
        if field.type == "struct" and field.fields:
            result.extend(_collect_leaf_fields(field.fields, col))
            col += field.column_span()
        else:
            result.append((field, col))
            col += 1
    return result


# ---------------------------------------------------------------------------
# Type annotation string
# ---------------------------------------------------------------------------

def _type_annotation(field: FieldDef) -> str:
    """Build a human-readable type annotation for the type row."""
    return TYPE_TRAITS[field.type].excel_annotation(field)


# ---------------------------------------------------------------------------
# Rich-text builder for the merged name + type cell
# ---------------------------------------------------------------------------

def _make_name_type_richtext(name: str, type_text: str) -> CellRichText:
    """Two-line rich text: name (bold white) on top, type (italic light) below.

    The newline is embedded at the end of the name run. A standalone
    whitespace-only ``TextBlock`` would be serialized as ``<t>\\n</t>`` without
    ``xml:space="preserve"``, which Excel strips per XML default whitespace
    handling — the result would render on a single line.
    """
    return CellRichText([
        TextBlock(_NAME_RUN_FONT, f"{name}\n"),
        TextBlock(_TYPE_RUN_FONT, type_text),
    ])


def _struct_type_label(field: FieldDef) -> str:
    """Type label shown beneath a struct field's name (matches FBS table name)."""
    return TYPE_TRAITS["struct"].excel_annotation(field)


# ---------------------------------------------------------------------------
# Recursive header writer
# ---------------------------------------------------------------------------

def _write_field_headers(
    ws,
    fields: list[FieldDef],
    start_col: int,
    group_row_start: int,
    group_row_end: int,
    comment_row: int,
    primary_key: str = "",
) -> int:
    """Write header cells for *fields* starting at *start_col*.

    Returns the next available column index (1-based).

    Parameters
    ----------
    ws : Worksheet
    fields : list of field definitions at this nesting level
    start_col : 1-based column index to start writing
    group_row_start : 1-based row for this nesting level's name+type cell
    group_row_end : 1-based row of the last name+type row (before comment)
    comment_row : 1-based row index for comments
    """
    col = start_col
    for field in fields:
        span = field.column_span()

        if field.type == "struct" and field.fields:
            # ---- struct: horizontal merge in current group row ----
            end_col = col + span - 1
            if span > 1:
                ws.merge_cells(
                    start_row=group_row_start, start_column=col,
                    end_row=group_row_start, end_column=end_col,
                )
            cell = ws.cell(row=group_row_start, column=col)
            cell.value = _make_name_type_richtext(field.name, _struct_type_label(field))
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.fill = _STRUCT_FILL
            cell.border = _THIN_BORDER

            for c in range(col, end_col + 1):
                ws.cell(row=group_row_start, column=c).border = _THIN_BORDER

            # Recurse into sub-fields one row deeper
            _write_field_headers(
                ws, field.fields,
                start_col=col,
                group_row_start=group_row_start + 1,
                group_row_end=group_row_end,
                comment_row=comment_row,
                primary_key=primary_key,
            )
            col = end_col + 1

        else:
            # ---- leaf field: vertical merge across remaining group rows ----
            is_primary = field.name == primary_key
            name_fill = _PRIMARY_FILL if is_primary else _NORMAL_FILL

            if group_row_start < group_row_end:
                ws.merge_cells(
                    start_row=group_row_start, start_column=col,
                    end_row=group_row_end, end_column=col,
                )
            cell = ws.cell(row=group_row_start, column=col)
            cell.value = _make_name_type_richtext(field.name, _type_annotation(field))
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.fill = name_fill
            cell.border = _THIN_BORDER

            for r in range(group_row_start, group_row_end + 1):
                ws.cell(row=r, column=col).border = _THIN_BORDER

            # Comment row
            comment_cell = ws.cell(row=comment_row, column=col,
                                   value=field.comment or "")
            comment_cell.font = _COMMENT_FONT
            comment_cell.alignment = _CENTER
            comment_cell.fill = _COMMENT_FILL
            comment_cell.border = _THIN_BORDER

            col += 1

    return col


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

# Metadata field names stored in Excel Custom Document Properties.
_META_TOOL_VERSION = "ct_tool_version"
_META_TABLE_NAME = "ct_table_name"
_META_HEADER_ROWS = "ct_header_rows"
_META_SCHEMA_HASH = "ct_schema_hash"
_META_GENERATED_AT = "ct_generated_at"

_META_REQUIRED = (
    _META_TOOL_VERSION,
    _META_TABLE_NAME,
    _META_HEADER_ROWS,
    _META_SCHEMA_HASH,
    _META_GENERATED_AT,
)


@dataclass(frozen=True)
class TemplateMetadata:
    """Self-describing metadata embedded in a generated Excel template."""

    tool_version: str
    table_name: str
    header_rows: int
    schema_hash: str
    generated_at: str


def _tool_version() -> str:
    try:
        return importlib_metadata.version("ct-tool")
    except importlib_metadata.PackageNotFoundError:
        return "unknown"


def _write_metadata(wb: Workbook, schema: TableSchema) -> None:
    """Write the five ct_* properties to the workbook's custom doc props."""
    props = wb.custom_doc_props
    # Drop any pre-existing ct_* entries so we never end up with duplicates.
    for name in _META_REQUIRED:
        try:
            del props[name]
        except (KeyError, AttributeError):
            pass

    now = datetime.now(timezone.utc)
    props.append(StringProperty(name=_META_TOOL_VERSION, value=_tool_version()))
    props.append(StringProperty(name=_META_TABLE_NAME, value=schema.table))
    props.append(IntProperty(name=_META_HEADER_ROWS, value=schema.header_rows))
    props.append(StringProperty(name=_META_SCHEMA_HASH, value=compute_schema_hash(schema)))
    props.append(DateTimeProperty(name=_META_GENERATED_AT, value=now))


def read_template_metadata(path: Path) -> TemplateMetadata | None:
    """Read ct_* metadata from an existing Excel file.

    Returns ``None`` for any failure mode: file missing, file corrupted,
    metadata missing, partial metadata, or unexpected types. The caller
    should treat ``None`` as "untracked / legacy" and fall back accordingly.
    """
    if not path.exists():
        return None
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        props = wb.custom_doc_props
        values: dict[str, object] = {}
        for prop in props:
            if prop.name in _META_REQUIRED:
                values[prop.name] = prop.value
        if not all(name in values for name in _META_REQUIRED):
            return None
        try:
            header_rows = int(values[_META_HEADER_ROWS])  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None
        generated_at = values[_META_GENERATED_AT]
        if isinstance(generated_at, datetime):
            generated_at_str = generated_at.isoformat()
        else:
            generated_at_str = str(generated_at)
        return TemplateMetadata(
            tool_version=str(values[_META_TOOL_VERSION]),
            table_name=str(values[_META_TABLE_NAME]),
            header_rows=header_rows,
            schema_hash=str(values[_META_SCHEMA_HASH]),
            generated_at=generated_at_str,
        )
    except Exception:
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def iter_data_rows(path: Path, header_rows: int) -> Iterator[tuple]:
    """产出表头下方的非空数据行（openpyxl values-only 元组）。

    "非空"定义为任一单元格 `not None`（与 `_has_data_rows` /
    `update_template` 的判断一致，不是字符串 strip 语义）。工作簿以
    只读模式打开，迭代结束后关闭。
    """
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx <= header_rows:
                continue
            if any(cell is not None for cell in row):
                yield row
    finally:
        wb.close()


def generate_template(schema: TableSchema, output_path: Path) -> None:
    """Create an Excel template workbook with structured headers."""
    total_rows = schema.header_rows  # max_nesting_depth + 1
    group_rows = schema.max_nesting_depth
    comment_row = group_rows + 1

    wb = Workbook()
    ws = wb.active
    ws.title = schema.table  # type: ignore[union-attr]

    _write_field_headers(
        ws,
        schema.fields,
        start_col=1,
        group_row_start=1,
        group_row_end=group_rows,
        comment_row=comment_row,
        primary_key=schema.primary,
    )

    total_cols = sum(f.column_span() for f in schema.fields)
    last_col_letter = get_column_letter(total_cols)

    # Fixed column width
    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16  # type: ignore[union-attr]

    # Explicit row height for the name+type rows so the two-line rich text
    # is fully visible on viewers that ignore wrap_text auto-fit.
    for r in range(1, group_rows + 1):
        ws.row_dimensions[r].height = _NAME_ROW_HEIGHT  # type: ignore[union-attr]

    # Freeze panes below the header
    ws.freeze_panes = ws.cell(row=total_rows + 1, column=1)  # type: ignore[union-attr]

    # DataValidation for enum fields
    data_start = total_rows + 1
    for field, col_idx in _collect_leaf_fields(schema.fields, 1):
        if field.type != "enum" or not field.values:
            continue
        formula_str = '"' + ",".join(field.values) + '"'
        if len(formula_str) > 255:
            logger.warning(
                "enum 字段 '%s' 的值列表超过 255 字符，跳过下拉菜单", field.name
            )
            continue
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=formula_str,
            showDropDown=False,
            allow_blank=True,
        )
        dv.sqref = f"{col_letter}{data_start}:{col_letter}1000"
        ws.add_data_validation(dv)  # type: ignore[union-attr]

    # Zebra striping via conditional formatting (fill + column separator borders)
    data_range = f"A{data_start}:{last_col_letter}1000"
    ws.conditional_formatting.add(  # type: ignore[union-attr]
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=_ZEBRA_FILL, border=_COL_BORDER),
    )
    ws.conditional_formatting.add(  # type: ignore[union-attr]
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=1"], fill=_WHITE_FILL, border=_COL_BORDER),
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_metadata(wb, schema)
    wb.save(str(output_path))
    logger.info("模板已生成: %s (%d 列, %d 行表头)",
                output_path, total_cols, total_rows)


def update_template(schema: TableSchema, output_path: Path) -> int:
    """Rebuild the header rows of an existing template, preserving data rows.

    Reads the file's metadata to determine where the old header ended, captures
    every non-empty row below it, regenerates the template (which writes fresh
    metadata), and appends the captured rows back below the new header.

    Returns the number of data rows that were preserved.

    For legacy files (no metadata) the current schema's ``header_rows`` is used
    as the best guess for where the old header ended — if the old schema's
    nesting depth differed, the caller is expected to warn the user.
    """
    meta = read_template_metadata(output_path)
    old_header_rows = meta.header_rows if meta is not None else schema.header_rows

    # 1. Snapshot data rows from the existing file.
    data_rows = list(iter_data_rows(output_path, old_header_rows))

    # 2. Regenerate the template with fresh headers + metadata.
    generate_template(schema, output_path)

    # 3. Append preserved data rows under the new header.
    if data_rows:
        new_wb = load_workbook(str(output_path))
        try:
            new_ws = new_wb.active
            for row in data_rows:
                new_ws.append(list(row))
            new_wb.save(str(output_path))
        finally:
            new_wb.close()

    logger.info(
        "模板已更新: %s (保留 %d 行数据，旧表头 %d 行)",
        output_path, len(data_rows), old_header_rows,
    )
    return len(data_rows)
