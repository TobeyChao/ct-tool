"""Integration tests for `ct status` template-drift detection.

Builds a minimal project root in tmp_path with config/global.yaml,
config/schemas/*.yaml, and excel/ files, then invokes the typer CLI.

Spec scenarios covered:
- Show pending data changes
- Show drifted templates
- Show untracked templates
- All clean reports nothing pending
"""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook
from typer.testing import CliRunner

from ct.cli import app
from ct.excel.template import generate_template
from ct.schema.models import FieldDef, TableSchema


runner = CliRunner()


def _setup_project(tmp_path: Path, schemas: dict[str, str]) -> Path:
    """Create a minimal project root with config + schemas. Returns project root."""
    (tmp_path / "config").mkdir()
    (tmp_path / "config" / "schemas").mkdir()
    (tmp_path / "excel").mkdir()
    (tmp_path / "cache").mkdir()
    (tmp_path / "i18n").mkdir()
    (tmp_path / "output").mkdir()

    (tmp_path / "config" / "global.yaml").write_text(
        yaml.safe_dump({"primary_lang": "zh"}),
        encoding="utf-8",
    )

    for name, body in schemas.items():
        (tmp_path / "config" / "schemas" / f"{name}.yaml").write_text(body, encoding="utf-8")

    return tmp_path


_ITEM_SCHEMA_YAML = """
table: item
primary: id
fields:
  - {name: id, type: int32, comment: 主键}
  - {name: name, type: string, comment: 名称}
  - {name: price, type: int32, comment: 价格}
"""

_ITEM_SCHEMA_YAML_V2 = """
table: item
primary: id
fields:
  - {name: id, type: int32, comment: 主键}
  - {name: name, type: string, comment: 名称}
  - {name: price, type: int32, comment: 价格}
  - {name: rarity, type: enum, values: [common, rare]}
"""


def _build_schema_from_yaml(yaml_text: str) -> TableSchema:
    data = yaml.safe_load(yaml_text)
    return TableSchema.model_validate(data)


def test_all_clean_reports_nothing_pending(tmp_path: Path) -> None:
    root = _setup_project(tmp_path, {"item": _ITEM_SCHEMA_YAML})
    schema = _build_schema_from_yaml(_ITEM_SCHEMA_YAML)
    xlsx = root / "excel" / "item.xlsx"
    generate_template(schema, xlsx)

    # Seed cache so the file's current hash is "known" — otherwise it shows as [changed].
    from ct.cache.state import CacheState, save_cache, update_table_cache
    from ct.excel.diff import file_hash

    cache = CacheState()
    update_table_cache(cache, "item", hash=file_hash(xlsx), ids=[])
    save_cache(cache, root / "cache")

    result = runner.invoke(app, ["status", "--root", str(root)])
    assert result.exit_code == 0, result.stdout
    assert "[OK] 所有表已是最新" in result.stdout


def test_drifted_template_is_reported(tmp_path: Path) -> None:
    root = _setup_project(tmp_path, {"item": _ITEM_SCHEMA_YAML})
    # Generate template with v1 schema...
    schema_v1 = _build_schema_from_yaml(_ITEM_SCHEMA_YAML)
    generate_template(schema_v1, root / "excel" / "item.xlsx")
    # ...then upgrade the schema yaml to v2 (mismatch with template metadata).
    (root / "config" / "schemas" / "item.yaml").write_text(_ITEM_SCHEMA_YAML_V2, encoding="utf-8")

    result = runner.invoke(app, ["status", "--root", str(root)])
    assert result.exit_code == 0, result.stdout
    assert "[template-stale] item" in result.stdout
    assert "ct gen-template --table item --update-header" in result.stdout


def test_untracked_template_is_reported(tmp_path: Path) -> None:
    root = _setup_project(tmp_path, {"item": _ITEM_SCHEMA_YAML})
    # Build a workbook with NO ct_* metadata.
    wb = Workbook()
    wb.active.title = "item"  # type: ignore[union-attr]
    wb.save(str(root / "excel" / "item.xlsx"))

    result = runner.invoke(app, ["status", "--root", str(root)])
    assert result.exit_code == 0, result.stdout
    assert "[template-untracked] item" in result.stdout


def test_data_change_is_reported(tmp_path: Path) -> None:
    """If excel changes after a baseline cache, status shows [changed]."""
    root = _setup_project(tmp_path, {"item": _ITEM_SCHEMA_YAML})
    schema = _build_schema_from_yaml(_ITEM_SCHEMA_YAML)
    xlsx = root / "excel" / "item.xlsx"
    generate_template(schema, xlsx)

    # Seed cache with the current file hash so a subsequent edit is "changed".
    from ct.cache.state import CacheState, save_cache, update_table_cache
    from ct.excel.diff import file_hash

    cache = CacheState()
    update_table_cache(cache, "item", hash=file_hash(xlsx), ids=[])
    save_cache(cache, root / "cache")

    # Mutate the file to bump its hash.
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx))
    wb.active.cell(row=schema.header_rows + 1, column=1, value=999)
    wb.save(str(xlsx))

    result = runner.invoke(app, ["status", "--root", str(root)])
    assert result.exit_code == 0, result.stdout
    assert "[changed] item" in result.stdout
