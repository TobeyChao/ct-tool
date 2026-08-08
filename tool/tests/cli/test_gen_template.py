"""Tests for ct.cli_helpers.template_action.decide_template_action.

Covers the 9 spec scenarios in cli-interface for ct gen-template:
- Generate template for all tables
- New file generates with metadata
- Hash matches default skips with hint
- Hash matches with --force rebuilds
- Hash differs with data refuses by default
- Hash differs with --update-header preserves data
- Legacy file without metadata refuses by default
- Legacy file with --update-header uses current schema header_rows
- Table name mismatch refuses even with --force
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.packaging.custom import StringProperty

from ct.cli_helpers.template_action import (
    Action,
    decide_template_action,
)
from ct.excel.template import generate_template
from ct.schema.models import FieldDef, TableSchema


def _schema(table: str = "Item") -> TableSchema:
    return TableSchema(
        table=table,
        primary="Id",
        fields=[
            FieldDef(name="Id", type="int32", comment="主键"),
            FieldDef(name="Name", type="string", comment="名称"),
            FieldDef(name="Price", type="int32", comment="价格"),
        ],
    )


def _schema_with_extra_field(table: str = "Item") -> TableSchema:
    s = _schema(table)
    s.fields.append(FieldDef(name="Rarity", type="enum", values=["common", "rare"]))
    return s


def _fill_data(path: Path, header_rows: int, rows: list[tuple]) -> None:
    wb = load_workbook(str(path))
    ws = wb.active
    for r_offset, row in enumerate(rows):
        for c_offset, value in enumerate(row):
            ws.cell(row=header_rows + 1 + r_offset, column=c_offset + 1, value=value)
    wb.save(str(path))


def _make_legacy_file(path: Path, table: str = "Item") -> None:
    """Build a workbook with no ct_* metadata."""
    wb = Workbook()
    wb.active.title = table  # type: ignore[union-attr]
    wb.save(str(path))


# -- Scenario: New file → CREATE_NEW --------------------------------------

def test_missing_file_creates_new(tmp_path: Path) -> None:
    schema = _schema()
    decision = decide_template_action(
        schema, tmp_path / "nope.xlsx", force=False, update_header=False,
    )
    assert decision.action == Action.CREATE_NEW
    assert "[new]" in decision.message


# -- Scenario: Hash matches → SKIP ----------------------------------------

def test_hash_matches_default_skips(tmp_path: Path) -> None:
    schema = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema, out)

    decision = decide_template_action(schema, out, force=False, update_header=False)
    assert decision.action == Action.SKIP
    assert "schema 未变化" in decision.message
    assert "--force" in decision.message


def test_hash_matches_with_force_rebuilds(tmp_path: Path) -> None:
    schema = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema, out)

    decision = decide_template_action(schema, out, force=True, update_header=False)
    assert decision.action == Action.REBUILD


# -- Scenario: Hash differs + data → REFUSE / UPDATE_PRESERVE / REBUILD ---

def test_hash_differs_with_data_refuses_by_default(tmp_path: Path) -> None:
    schema_v1 = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema_v1, out)
    _fill_data(out, schema_v1.header_rows, [(1, "sword", 100)])

    schema_v2 = _schema_with_extra_field()
    decision = decide_template_action(schema_v2, out, force=False, update_header=False)
    assert decision.action == Action.REFUSE
    assert "--update-header" in decision.message
    assert "--force" in decision.message


def test_hash_differs_with_data_and_update_header_preserves(tmp_path: Path) -> None:
    schema_v1 = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema_v1, out)
    _fill_data(out, schema_v1.header_rows, [(1, "sword", 100)])

    schema_v2 = _schema_with_extra_field()
    decision = decide_template_action(schema_v2, out, force=False, update_header=True)
    assert decision.action == Action.UPDATE_PRESERVE


def test_hash_differs_with_data_and_force_rebuilds(tmp_path: Path) -> None:
    schema_v1 = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema_v1, out)
    _fill_data(out, schema_v1.header_rows, [(1, "sword", 100)])

    schema_v2 = _schema_with_extra_field()
    decision = decide_template_action(schema_v2, out, force=True, update_header=False)
    assert decision.action == Action.REBUILD


def test_hash_differs_no_data_rebuilds_silently(tmp_path: Path) -> None:
    """Spec: hash differs + no data → directly rebuild (no flag needed)."""
    schema_v1 = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema_v1, out)

    schema_v2 = _schema_with_extra_field()
    decision = decide_template_action(schema_v2, out, force=False, update_header=False)
    assert decision.action == Action.REBUILD


# -- Scenario: Legacy file (no metadata) ----------------------------------

def test_legacy_file_refuses_by_default(tmp_path: Path) -> None:
    out = tmp_path / "legacy.xlsx"
    _make_legacy_file(out)

    decision = decide_template_action(_schema(), out, force=False, update_header=False)
    assert decision.action == Action.REFUSE
    assert "无元数据" in decision.message


def test_legacy_file_update_header_preserves(tmp_path: Path) -> None:
    out = tmp_path / "legacy.xlsx"
    _make_legacy_file(out)

    decision = decide_template_action(_schema(), out, force=False, update_header=True)
    assert decision.action == Action.UPDATE_PRESERVE
    assert "legacy" in decision.message.lower() or "无元数据" in decision.message


def test_legacy_file_force_rebuilds(tmp_path: Path) -> None:
    out = tmp_path / "legacy.xlsx"
    _make_legacy_file(out)

    decision = decide_template_action(_schema(), out, force=True, update_header=False)
    assert decision.action == Action.REBUILD


# -- Scenario: Table name mismatch ----------------------------------------

def test_table_name_mismatch_refuses_default(tmp_path: Path) -> None:
    out = tmp_path / "item.xlsx"
    generate_template(_schema("Quest"), out)  # File belongs to 'quest'

    decision = decide_template_action(
        _schema("Item"), out, force=False, update_header=False,
    )
    assert decision.action == Action.REFUSE
    assert "Quest" in decision.message
    assert "Item" in decision.message


def test_table_name_mismatch_refuses_even_with_force(tmp_path: Path) -> None:
    out = tmp_path / "item.xlsx"
    generate_template(_schema("Quest"), out)

    decision = decide_template_action(
        _schema("Item"), out, force=True, update_header=False,
    )
    assert decision.action == Action.REFUSE


def test_table_name_mismatch_refuses_even_with_update_header(tmp_path: Path) -> None:
    out = tmp_path / "item.xlsx"
    generate_template(_schema("Quest"), out)

    decision = decide_template_action(
        _schema("Item"), out, force=False, update_header=True,
    )
    assert decision.action == Action.REFUSE
