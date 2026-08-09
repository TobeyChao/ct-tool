"""Read Excel data files using openpyxl read_only mode and parse rows
according to a TableSchema definition.

Key concepts:
- Header rows are determined by ``schema.header_rows`` (max_nesting_depth + 1).
- Struct fields are expanded into multiple contiguous columns.  A struct with
  two sub-fields occupies two columns; nested structs expand recursively.
- Array fields occupy a single cell whose value is split by
  ``field.separator``.
- Empty rows (all *None*) are silently skipped.
- Extra columns beyond those declared in the schema trigger a warning.
"""

from __future__ import annotations

from dataclasses import dataclass
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ct.schema.models import FieldDef, TableSchema

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Column span helpers
# ---------------------------------------------------------------------------

def _column_span(field: FieldDef) -> int:
    """Return how many Excel columns *field* occupies.

    - A basic / enum / array field occupies exactly 1 column.
    - A struct field occupies the sum of its sub-fields' spans (recursive).
    """
    if field.type == "struct" and field.fields:
        return sum(_column_span(sf) for sf in field.fields)
    return 1


def _flatten_fields(fields: list[FieldDef]) -> list[tuple[str, FieldDef]]:
    """Return a flat list of ``(dotted_path, leaf_field)`` tuples.

    For example, a struct ``drop_range`` with sub-fields ``min`` and ``max``
    yields::

        [("drop_range.min", <FieldDef min>),
         ("drop_range.max", <FieldDef max>)]

    Non-struct fields simply yield ``(name, field)``.
    """
    result: list[tuple[str, FieldDef]] = []
    for f in fields:
        if f.type == "struct" and f.fields:
            for sub_path, sub_field in _flatten_fields(f.fields):
                result.append((f"{f.name}.{sub_path}", sub_field))
        else:
            result.append((f.name, f))
    return result


def leaf_column_map(schema: TableSchema) -> dict[str, int]:
    """返回 ``{dotted_path: 0-based 列索引}``（struct 展开后按叶子列计）。"""
    result: dict[str, int] = {}
    col = 0
    for path, leaf in _flatten_fields(schema.fields):
        result[path] = col
        col += _column_span(leaf)
    return result


# ---------------------------------------------------------------------------
# Type coercion
# ---------------------------------------------------------------------------

_BOOL_TRUE = frozenset({"true", "1", "yes", "TRUE", "True", "YES", "Yes"})
_BOOL_FALSE = frozenset({"false", "0", "no", "FALSE", "False", "NO", "No"})


def _coerce(value: Any, field: FieldDef) -> Any:
    """Convert a raw cell value to the expected Python type."""
    if value is None:
        return None

    type_name = field.type
    # For leaf fields inside structs, type_name is already the leaf type.

    if type_name in ("int32", "int64"):
        if isinstance(value, float) and value == int(value):
            return int(value)
        return int(value)
    if type_name in ("float", "double"):
        return float(value)
    if type_name == "bool":
        if isinstance(value, bool):
            return value
        s = str(value).strip()
        if s in _BOOL_TRUE:
            return True
        if s in _BOOL_FALSE:
            return False
        raise ValueError(f"无法将 '{value}' 转换为 bool")
    if type_name == "string":
        return str(value)
    if type_name == "enum":
        return str(value)
    # array and struct are handled separately
    return value


def _coerce_element(value: str, element_type: str) -> Any:
    """Coerce a single array element string to the declared element type."""
    value = value.strip()
    if element_type in ("int32", "int64"):
        return int(float(value)) if "." in value else int(value)
    if element_type in ("float", "double"):
        return float(value)
    if element_type == "bool":
        if value in _BOOL_TRUE:
            return True
        if value in _BOOL_FALSE:
            return False
        raise ValueError(f"无法将数组元素 '{value}' 转换为 bool")
    # string / enum
    return value


# ---------------------------------------------------------------------------
# Row parsing
# ---------------------------------------------------------------------------

def _build_nested_dict(flat: dict[str, Any]) -> dict[str, Any]:
    """Convert a flat dict with dotted keys into a nested dict.

    >>> _build_nested_dict({"drop_range.min": 10, "drop_range.max": 20})
    {"drop_range": {"min": 10, "max": 20}}
    """
    result: dict[str, Any] = {}
    for dotted_key, val in flat.items():
        parts = dotted_key.split(".")
        d = result
        for part in parts[:-1]:
            d = d.setdefault(part, {})
        d[parts[-1]] = val
    return result


def _parse_row(
    cells: tuple[Any, ...],
    flat_columns: list[tuple[str, FieldDef]],
    top_fields: list[FieldDef],
) -> dict[str, Any] | None:
    """Parse a single data row into a dict.  Returns *None* if the row is
    entirely empty."""
    # Check if the row is empty (all None / empty string)
    if all(c is None or (isinstance(c, str) and not c.strip()) for c in cells):
        return None

    flat: dict[str, Any] = {}
    for col_idx, (dotted_path, field) in enumerate(flat_columns):
        if col_idx >= len(cells):
            raw = None
        else:
            raw = cells[col_idx]

        # Determine if this is an array field (arrays are always top-level,
        # never inside structs, per model validation).
        # We check against the top-level field name.
        top_name = dotted_path.split(".")[0]
        top_field = next((f for f in top_fields if f.name == top_name), None)

        if top_field and top_field.type == "array" and "." not in dotted_path:
            # Array: split cell by separator
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                flat[dotted_path] = []
            else:
                raw_str = str(raw)
                sep = top_field.separator or ","
                elements = [
                    _coerce_element(e, top_field.element)
                    for e in raw_str.split(sep)
                    if e.strip()
                ]
                flat[dotted_path] = elements
        else:
            flat[dotted_path] = _coerce(raw, field)

    return _build_nested_dict(flat)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ParsedRows:
    """解析结果：数据行 + 与之一一对应的 Excel 绝对行号。

    数据与定位信息分离，避免行号泄漏进 JSON / Binary 等导出产物。
    """

    rows: list[dict[str, Any]]
    excel_rows: list[int]


def read_excel(excel_path: Path, schema: TableSchema) -> ParsedRows:
    """Read an Excel file and return parsed data rows.

    Parameters
    ----------
    excel_path:
        Path to the ``.xlsx`` file.
    schema:
        The ``TableSchema`` describing the table structure.

    Returns
    -------
    ParsedRows
        ``rows``: 与旧返回完全一致的数据行（struct 为嵌套 dict）；
        ``excel_rows``: 与 ``rows`` 平行的 Excel 绝对行号（空行被跳过，
        但真实行号保留）。
    """
    wb = load_workbook(str(excel_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            logger.warning("工作簿 %s 没有活动工作表", excel_path)
            return ParsedRows([], [])

        header_row_count = schema.header_rows
        flat_columns = _flatten_fields(schema.fields)
        expected_col_count = len(flat_columns)

        rows: list[dict[str, Any]] = []
        excel_rows: list[int] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Skip header rows
            if row_idx <= header_row_count:
                # On the last header row, check for extra columns
                if row_idx == header_row_count:
                    actual_cols = len(row)
                    if actual_cols > expected_col_count:
                        extra = actual_cols - expected_col_count
                        logger.warning(
                            "表 %s (%s): 发现 %d 个多余列（schema 定义 %d 列，"
                            "Excel 有 %d 列）",
                            schema.table, excel_path.name,
                            extra, expected_col_count, actual_cols,
                        )
                continue

            parsed = _parse_row(row, flat_columns, schema.fields)
            if parsed is not None:
                rows.append(parsed)
                excel_rows.append(row_idx)

        return ParsedRows(rows, excel_rows)
    finally:
        wb.close()
