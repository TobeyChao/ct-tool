"""Canonical template workbook generation tests (4.2)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ct.excel.canonical_template import generate_canonical_template
from ct.excel.layout import build_layout
from ct.schema.resources import EnumResource, FieldDef, RecordResource, TableResource


def _table() -> TableResource:
    return TableResource(
        table="Item",
        primary="Id",
        fields=[
            FieldDef(name="Id", type="int32", comment="主键"),
            FieldDef(name="Rarity", type="ItemRarity", comment="品质"),
            FieldDef(name="DropRange", type="DropRange", comment="掉落区间"),
            FieldDef(
                name="Rewards",
                type="vector<DropReward>",
                excel_columns=2,
                comment="奖励组",
            ),
        ],
    )


def _records() -> dict[str, RecordResource]:
    return {
        "DropRange": RecordResource(
            name="DropRange",
            fields=[
                FieldDef(name="Min", type="int32", comment="下限"),
                FieldDef(name="Max", type="int32", comment="上限"),
            ],
        ),
        "DropReward": RecordResource(
            name="DropReward",
            fields=[
                FieldDef(name="ItemId", type="int32", comment="道具"),
                FieldDef(name="Count", type="int32", comment="数量"),
            ],
        ),
    }


def test_template_headers_follow_layout(tmp_path: Path) -> None:
    table = _table()
    records = _records()
    layout = build_layout(table, schema_hash="sha1", records=records)
    enums = {"ItemRarity": EnumResource(name="ItemRarity", values=["Common", "Rare"])}
    out = generate_canonical_template(
        layout, tmp_path / "item.xlsx", enums=enums, primary=table.primary
    )

    assert out.exists()
    wb = load_workbook(str(out))
    ws = wb.active
    assert ws.max_row == layout.header_rows
    # row 1 top-level fields: Id(1) Rarity(2) DropRange(3-4) Rewards(5-8)
    assert "Rewards" in str(ws.cell(row=1, column=5).value)
    assert "vector<DropReward>" in str(ws.cell(row=1, column=5).value)
    assert "DropRange" in str(ws.cell(row=1, column=3).value)
    # group header on row 2
    assert "1" in str(ws.cell(row=2, column=5).value)
    # record leaf on row 2 (DropRange depth 2)
    assert "Min" in str(ws.cell(row=2, column=3).value)
    # comment row: per-leaf comments
    assert ws.cell(row=layout.header_rows, column=1).value == "主键"
    assert ws.cell(row=layout.header_rows, column=3).value == "下限"
    assert ws.cell(row=layout.header_rows, column=5).value == "道具"

    # metadata
    props = {prop.name: prop.value for prop in wb.custom_doc_props}
    assert props["ct_schema_hash"] == "sha1"
    assert props["ct_header_rows"] == layout.header_rows

    # enum dropdown on Rarity column (col 2)
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("Common" in formula and "Rare" in formula for formula in formulas)
    wb.close()


def test_template_golden_stable_across_runs(tmp_path: Path) -> None:
    table = _table()
    records = _records()
    layout = build_layout(table, schema_hash="sha2", records=records)
    enums = {"ItemRarity": EnumResource(name="ItemRarity", values=["Common", "Rare"])}
    first = tmp_path / "a.xlsx"
    second = tmp_path / "b.xlsx"
    generate_canonical_template(layout, first, enums=enums, primary=table.primary)
    generate_canonical_template(layout, second, enums=enums, primary=table.primary)
    assert first.read_bytes() == second.read_bytes()
