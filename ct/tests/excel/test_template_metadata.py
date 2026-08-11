"""Tests for ct.excel.template metadata read/write."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from ct.excel.template import (
    TemplateMetadata,
    generate_template,
    read_template_metadata,
)
from ct.schema.models import FieldDef, TableSchema


def _schema() -> TableSchema:
    return TableSchema(
        table="Item",
        primary="Id",
        fields=[
            FieldDef(name="Id", type="int32", comment="主键"),
            FieldDef(name="Name", type="string", i18n=True, comment="名称"),
            FieldDef(name="Price", type="int32", comment="价格"),
        ],
    )


def test_generate_template_writes_all_metadata(tmp_path: Path) -> None:
    schema = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema, out)

    meta = read_template_metadata(out)
    assert meta is not None
    assert isinstance(meta, TemplateMetadata)
    assert meta.table_name == "Item"
    assert meta.header_rows == schema.header_rows
    assert len(meta.schema_hash) == 16
    assert meta.tool_version  # non-empty
    assert meta.generated_at  # non-empty


def test_metadata_invisible_in_sheet(tmp_path: Path) -> None:
    """ct_* keys must not appear as cell values, sheet names, or defined names."""
    schema = _schema()
    out = tmp_path / "item.xlsx"
    generate_template(schema, out)

    wb = load_workbook(str(out))
    assert wb.sheetnames == ["Item"]
    assert list(wb.defined_names) == []
    ws = wb.active
    visible_cells = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                visible_cells.append(str(cell))
    blob = "\n".join(visible_cells)
    for forbidden in ("ct_table_name", "ct_schema_hash", "ct_header_rows"):
        assert forbidden not in blob


def test_read_returns_none_for_missing_file(tmp_path: Path) -> None:
    assert read_template_metadata(tmp_path / "nope.xlsx") is None


def test_read_returns_none_for_file_without_metadata(tmp_path: Path) -> None:
    """A plain workbook with no ct_* props yields None."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.cell(row=1, column=1, value="hello")
    out = tmp_path / "blank.xlsx"
    wb.save(str(out))

    assert read_template_metadata(out) is None


def test_read_returns_none_for_partial_metadata(tmp_path: Path) -> None:
    """Missing any required ct_* property → None (untrusted)."""
    from openpyxl import Workbook
    from openpyxl.packaging.custom import StringProperty

    wb = Workbook()
    # Only write one prop, leave the rest off.
    wb.custom_doc_props.append(StringProperty(name="ct_table_name", value="Item"))
    out = tmp_path / "partial.xlsx"
    wb.save(str(out))

    assert read_template_metadata(out) is None


def test_read_returns_none_for_corrupted_file(tmp_path: Path) -> None:
    """Garbage bytes that openpyxl rejects must not crash the caller."""
    bogus = tmp_path / "broken.xlsx"
    bogus.write_bytes(b"not a real xlsx file")

    assert read_template_metadata(bogus) is None


def test_round_trip_after_regeneration(tmp_path: Path) -> None:
    """Regenerating into the same path overwrites old metadata cleanly."""
    out = tmp_path / "item.xlsx"
    generate_template(_schema(), out)
    first = read_template_metadata(out)

    schema2 = _schema()
    schema2.fields.append(FieldDef(name="Rarity", type="enum", values=["common", "rare"]))
    generate_template(schema2, out)
    second = read_template_metadata(out)

    assert first is not None and second is not None
    assert first.schema_hash != second.schema_hash
    assert second.header_rows == schema2.header_rows