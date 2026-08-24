from __future__ import annotations

import json
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook
from typer.testing import CliRunner

from ct.cli import app

runner = CliRunner()


def _build_project(root: Path) -> None:
    """构建一个最小项目，含 item 表（i18n=name），不依赖 flatc。"""
    (root / "config" / "schemas").mkdir(parents=True)
    (root / "excel").mkdir()
    (root / "i18n").mkdir()
    (root / "cache").mkdir()

    cfg = {
        "primary_lang": "zh",
        "secondary_langs": ["en"],
        "schemas_dir": "config/schemas",
        "excel_dir": "excel",
        "output_dir": "output",
        "cache_dir": "cache",
        "i18n_dir": "i18n",
    }
    (root / "config" / "global.yaml").write_text(
        yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8"
    )
    schema = {
        "table": "Item",
        "primary": "Id",
        "fields": [
            {"name": "Id", "type": "int32"},
            {"name": "Name", "type": "string", "i18n": True},
            {"name": "Price", "type": "float"},
        ],
    }
    (root / "config" / "schemas" / "Item.yaml").write_text(
        yaml.safe_dump(schema, allow_unicode=True), encoding="utf-8"
    )

    wb = Workbook()
    ws = wb.active
    ws.append(["id", "name", "price"])
    ws.append(["主键", "名称", "价格"])
    ws.append([1001, "铁剑", 100.0])
    wb.save(root / "excel" / "item.xlsx")


def _add_quest_table(root: Path) -> None:
    """追加一张带 i18n 字段的 Quest 表。"""
    schema = {
        "table": "Quest",
        "primary": "Id",
        "fields": [
            {"name": "Id", "type": "int32"},
            {"name": "Title", "type": "string", "i18n": True},
        ],
    }
    (root / "config" / "schemas" / "Quest.yaml").write_text(
        yaml.safe_dump(schema, allow_unicode=True), encoding="utf-8"
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "title"])
    ws.append(["主键", "标题"])
    ws.append([1, "新手任务"])
    wb.save(root / "excel" / "Quest.xlsx")


def _add_row_to_excel(root: Path, row: tuple) -> None:
    from openpyxl import load_workbook
    path = root / "excel" / "item.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    ws.append(list(row))
    wb.save(path)


def test_export_runs_internal_sync_creating_lang_files(tmp_path: Path) -> None:
    _build_project(tmp_path)
    result = runner.invoke(app, ["export", "--all", "--root", str(tmp_path)])
    assert result.exit_code == 0, result.output

    en = tmp_path / "i18n" / "en" / "Item.json"
    assert en.exists()
    data = json.loads(en.read_text(encoding="utf-8"))
    assert "1001.Name" in data
    assert data["1001.Name"]["status"] == "missing"


def test_export_after_new_row_adds_missing_entry(tmp_path: Path) -> None:
    _build_project(tmp_path)
    runner.invoke(app, ["export", "--all", "--root", str(tmp_path)])

    _add_row_to_excel(tmp_path, (1002, "魔杖", 200.0))
    result = runner.invoke(app, ["export", "--all", "--root", str(tmp_path)])
    assert result.exit_code == 0, result.output

    data = json.loads((tmp_path / "i18n" / "en" / "Item.json").read_text(encoding="utf-8"))
    assert "1002.Name" in data
    assert data["1002.Name"]["status"] == "missing"


def test_export_verbose_emits_sync_summary(tmp_path: Path) -> None:
    _build_project(tmp_path)
    result = runner.invoke(
        app, ["export", "--all", "--verbose", "--root", str(tmp_path)]
    )
    assert result.exit_code == 0, result.output
    assert "[i18n sync]" in result.output


def test_export_incremental_without_flags(tmp_path: Path) -> None:
    """无参数增量导出：首次按变更导出，缓存一致后跳过。"""
    _build_project(tmp_path)

    result = runner.invoke(app, ["export", "--root", str(tmp_path)])
    assert result.exit_code == 0, result.output
    assert "导出完成" in result.output

    result2 = runner.invoke(app, ["export", "--root", str(tmp_path)])
    assert result2.exit_code == 0, result2.output
    assert "所有表均无变化，跳过导出" in result2.output


def test_export_single_table_keeps_other_i18n_files(tmp_path: Path) -> None:
    """单表导出不得清理其他表的 lang 文件（既有缺陷回归测试）。"""
    _build_project(tmp_path)
    _add_quest_table(tmp_path)

    first = runner.invoke(app, ["export", "--all", "--root", str(tmp_path)])
    assert first.exit_code == 0, first.output
    assert (tmp_path / "i18n" / "en" / "Item.json").exists()
    assert (tmp_path / "i18n" / "en" / "Quest.json").exists()

    single = runner.invoke(app, ["export", "--table", "Item", "--root", str(tmp_path)])
    assert single.exit_code == 0, single.output
    assert (tmp_path / "i18n" / "en" / "Quest.json").exists()
    assert (tmp_path / "i18n" / "en" / "Item.json").exists()


def test_incremental_export_keeps_other_i18n_files(tmp_path: Path) -> None:
    """部分表变化的增量导出不得清理其他表的 lang 文件（既有缺陷回归）。"""
    _build_project(tmp_path)
    _add_quest_table(tmp_path)

    first = runner.invoke(app, ["export", "--all", "--root", str(tmp_path)])
    assert first.exit_code == 0, first.output
    assert (tmp_path / "i18n" / "en" / "Quest.json").exists()

    # 只改 Item，Quest 未变化
    from openpyxl import load_workbook

    wb = load_workbook(tmp_path / "excel" / "item.xlsx")
    wb.active.append([1002, "魔杖", 200.0])
    wb.save(tmp_path / "excel" / "item.xlsx")

    result = runner.invoke(app, ["export", "--root", str(tmp_path)])
    assert result.exit_code == 0, result.output
    assert (tmp_path / "i18n" / "en" / "Quest.json").exists()
    assert (tmp_path / "i18n" / "en" / "Item.json").exists()


def test_i18n_sync_table_filter_keeps_other_i18n_files(tmp_path: Path) -> None:
    """`i18n sync --table` 不得清理其他表的 lang 文件。"""
    _build_project(tmp_path)
    _add_quest_table(tmp_path)

    runner.invoke(app, ["i18n", "sync", "--root", str(tmp_path)])
    assert (tmp_path / "i18n" / "en" / "Quest.json").exists()

    result = runner.invoke(
        app, ["i18n", "sync", "--table", "Item", "--root", str(tmp_path)]
    )
    assert result.exit_code == 0, result.output
    assert (tmp_path / "i18n" / "en" / "Quest.json").exists()
    assert (tmp_path / "i18n" / "en" / "Item.json").exists()


def test_incremental_export_rebuilds_missing_cache_bytes(tmp_path: Path) -> None:
    """未变化表的 fbs_bytes 缓存缺失时，增量导出必须重建而非静默丢表。"""
    _build_project(tmp_path)
    _add_quest_table(tmp_path)

    first = runner.invoke(app, ["export", "--all", "--root", str(tmp_path)])
    assert first.exit_code == 0, first.output
    quest_cache = tmp_path / "cache" / "fbs_bytes" / "Quest.bin"
    assert quest_cache.exists()
    quest_cache.unlink()

    # 只改 Item，触发增量导出（Quest 按 cache hash 判定未变化）
    from openpyxl import load_workbook

    wb = load_workbook(tmp_path / "excel" / "item.xlsx")
    wb.active.append([1002, "魔杖", 200.0])
    wb.save(tmp_path / "excel" / "item.xlsx")

    result = runner.invoke(app, ["export", "--root", str(tmp_path)])
    assert result.exit_code == 0, result.output

    # Quest 缓存被重建，且 bundle 中确实包含 Quest 表数据
    assert quest_cache.exists()
    bundle = (tmp_path / "output" / "binary" / "data_zh.bin").read_bytes()
    assert b"Quest" in bundle
