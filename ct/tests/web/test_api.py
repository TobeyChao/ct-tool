"""Web 面板 API 集成测试（Flask test client）。"""

from __future__ import annotations

import json
import time
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook

from ct.web.app import create_app


def _build_project(root: Path) -> None:
    """构造最小项目：Item 表（i18n=Name），不依赖 flatc。"""
    (root / "config" / "schemas").mkdir(parents=True)
    (root / "excel").mkdir()
    (root / "i18n").mkdir()
    (root / "cache").mkdir()
    (root / "output").mkdir()
    (root / "config" / "global.yaml").write_text(
        yaml.safe_dump(
            {
                "primary_lang": "zh",
                "secondary_langs": ["en"],
            },
            allow_unicode=True,
        ),
        encoding="utf-8",
    )
    (root / "config" / "schemas" / "Item.yaml").write_text(
        yaml.safe_dump(
            {
                "table": "Item",
                "primary": "Id",
                "fields": [
                    {"name": "Id", "type": "int32"},
                    {"name": "Name", "type": "string", "i18n": True},
                    {"name": "Price", "type": "float"},
                ],
            },
            allow_unicode=True,
        ),
        encoding="utf-8",
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "name", "price"])
    ws.append(["主键", "名称", "价格"])
    ws.append([1001, "铁剑", 100.0])
    ws.append([1002, "法杖", 200.0])
    wb.save(root / "excel" / "item.xlsx")


def _write_excel_rows(root: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "name", "price"])
    ws.append(["主键", "名称", "价格"])
    for row in rows:
        ws.append(row)
    wb.save(root / "excel" / "item.xlsx")


def _wait_export(client, timeout: float = 20.0) -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = client.get("/api/export/progress")
        data = resp.get_json()["data"]
        if data["status"] in ("done", "error", "cancelled"):
            return data
        time.sleep(0.2)
    raise AssertionError("导出超时未完成")


def _make_client(tmp_path):
    root = tmp_path / "gd"
    _build_project(root)
    return create_app(root).test_client(), root


def test_workspace_ok(tmp_path):
    client, _ = _make_client(tmp_path)
    resp = client.get("/api/workspace")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["ok"] is True
    assert data["data"]["config"]["primary_lang"] == "zh"
    assert data["data"]["status"]["missing"] == []


def test_workspace_missing_config(tmp_path):
    root = tmp_path / "nope"
    client = create_app(root).test_client()
    resp = client.get("/api/workspace")
    assert resp.status_code == 400
    assert resp.get_json()["ok"] is False
    assert "工作区不可用" in resp.get_json()["error"]


def test_export_success_and_history(tmp_path):
    client, root = _make_client(tmp_path)
    resp = client.post("/api/export", json={"forced": True})
    assert resp.status_code == 200
    state = _wait_export(client)
    assert state["status"] == "done", state
    assert (root / "output" / "json" / "item_zh.json").exists()
    history = client.get("/api/history").get_json()["data"]
    assert len(history) == 1
    assert history[0]["result"] == "成功"


def test_export_validation_error_logged(tmp_path):
    client, root = _make_client(tmp_path)
    _write_excel_rows(root, [[1001, "铁剑", 100.0], [1001, "重复", 1.0]])
    client.post("/api/export", json={"forced": True})
    state = _wait_export(client)
    assert state["status"] == "error"
    assert state["errors"]
    logs = client.get("/api/logs?module=校验").get_json()["data"]
    assert any("Id" in r["message"] or "重复" in r["message"] for r in logs)


def test_export_cancel(tmp_path):
    client, _ = _make_client(tmp_path)
    client.post("/api/export", json={"forced": True})
    resp = client.post("/api/export/cancel")
    assert resp.status_code == 200
    state = _wait_export(client)
    assert state["status"] in ("cancelled", "done")


def test_i18n_sync_entries_save_compact(tmp_path):
    client, root = _make_client(tmp_path)
    tables = client.get("/api/i18n/tables").get_json()["data"]
    assert any(t["table"] == "Item" and t["has_i18n"] for t in tables)

    resp = client.post("/api/i18n/sync", json={"table": "Item"})
    assert resp.status_code == 200
    assert resp.get_json()["data"]["langs"]["en"]["missing"] == 2

    entries = client.get("/api/i18n/entries?table=Item&lang=en").get_json()["data"]
    assert len(entries) == 2
    assert all(e["status"] == "missing" for e in entries)

    key = "1001.Name"
    resp = client.post(
        "/api/i18n/entry",
        json={"table": "Item", "lang": "en", "key": key, "text": "Iron Sword", "confirmed": True},
    )
    assert resp.status_code == 200
    entries = client.get("/api/i18n/entries?table=Item&lang=en").get_json()["data"]
    by_key = {e["key"]: e for e in entries}
    assert by_key[key]["status"] == "translated"

    # 删一行再 sync → orphan，compact 清理
    _write_excel_rows(root, [[1001, "铁剑", 100.0]])
    client.post("/api/i18n/sync", json={"table": "Item"})
    preview = client.post(
        "/api/i18n/compact", json={"table": "Item", "dry_run": True}
    ).get_json()["data"]
    assert preview["touched"] is True
    assert preview["total_removed"] == 1
    result = client.post(
        "/api/i18n/compact", json={"table": "Item", "dry_run": False}
    ).get_json()["data"]
    assert result["total_removed"] == 1

    status = client.get("/api/i18n/status").get_json()["data"]
    assert "en" in status


def test_schema_legacy_entry_is_read_only(tmp_path):
    client, root = _make_client(tmp_path)
    # 列表与详情
    schemas = client.get("/api/schemas").get_json()["data"]
    assert [s["table"] for s in schemas] == ["Item"]
    detail = client.get("/api/schemas/Item").get_json()["data"]
    assert detail["primary"] == "Id"
    # 测试夹具的 Excel 是手写的（无模板元数据），属 legacy/未跟踪
    assert detail["template_status"] == "untracked"

    schema_path = root / "config" / "schemas" / "Item.yaml"
    excel_path = root / "excel" / "item.xlsx"
    original_schema = schema_path.read_bytes()
    original_excel = excel_path.read_bytes()

    # 旧写协议已退役：写路由不再存在（405），不能直接改 YAML/Excel。
    resp = client.post(
        "/api/schemas",
        json={
            "table": "Quest",
            "primary": "Id",
            "fields_yaml": "- name: Id\n  type: int32\n- name: Name\n  type: string\n  i18n: true",
        },
    )
    assert resp.status_code == 405
    assert not (root / "config" / "schemas" / "Quest.yaml").exists()
    assert not (root / "excel" / "quest.xlsx").exists()

    resp = client.put("/api/schemas/Item", json={})
    assert resp.status_code == 405

    resp = client.delete("/api/schemas/Item")
    assert resp.status_code == 405
    assert schema_path.read_bytes() == original_schema
    assert excel_path.read_bytes() == original_excel


def test_schema_write_routes_removed_after_cutover(tmp_path):
    root = tmp_path / "gd"
    _build_project(root)
    client = create_app(root).test_client()
    assert client.post("/api/schemas", json={}).status_code == 405
    assert client.put("/api/schemas/Item", json={}).status_code == 405
    assert client.delete("/api/schemas/Item").status_code == 405


def test_logs_module_filter(tmp_path):
    client, _ = _make_client(tmp_path)
    client.post("/api/export", json={"forced": True})
    _wait_export(client)
    logs = client.get("/api/logs?module=导出").get_json()["data"]
    assert any("步骤" in r["message"] or "导出" in r["message"] for r in logs)
