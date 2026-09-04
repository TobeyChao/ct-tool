"""canonical_status data-change / template-drift detection tests.

These lock in the P1 fix: ``status`` must report a table as ``changed`` when the
Excel file hash differs from the ledger recorded at the last successful export
(a pure data edit, no schema change), and as ``drifted`` when the layout
manifest's ``schema_hash`` no longer matches the current schema.
"""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook
from typer.testing import CliRunner

from ct.app.canonical_commands import canonical_status
from ct.cli import app

runner = CliRunner()


def _build_project(root: Path) -> None:
    """Minimal canonical workspace: one Item table with an Excel data file."""
    (root / "config" / "schemas").mkdir(parents=True)
    (root / "excel").mkdir()
    (root / "i18n").mkdir()
    (root / "cache").mkdir()
    cfg = {
        "primary_lang": "zh",
        "secondary_langs": ["en"],
        "schemas_dir": "config/schemas",
        "excel_dir": "excel",
        "output_dir": "output",
        "cache_dir": "cache",
        "i18n_dir": "i18n",
    }
    (root / "config" / "global.yaml").write_text(
        yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8"
    )
    schema = {
        "table": "Item",
        "primary": "Id",
        "fields": [
            {"name": "Id", "type": "int32"},
            {"name": "Name", "type": "string", "i18n": True},
            {"name": "Price", "type": "float"},
        ],
    }
    (root / "config" / "schemas" / "Item.yaml").write_text(
        yaml.safe_dump(schema, allow_unicode=True), encoding="utf-8"
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "name", "price"])
    ws.append(["主键", "名称", "价格"])
    ws.append([1001, "铁剑", 100.0])
    wb.save(root / "excel" / "Item.xlsx")


def _set_price(root: Path, price: float) -> None:
    wb = load_workbook(root / "excel" / "Item.xlsx")
    ws = wb.active
    ws.cell(row=3, column=3, value=price)
    wb.save(root / "excel" / "Item.xlsx")


def test_status_is_clean_after_export(tmp_path: Path) -> None:
    root = tmp_path / "gd"
    _build_project(root)
    result = runner.invoke(app, ["export", "--all", "--root", str(root)])
    assert result.exit_code == 0, result.output
    report = canonical_status(root)
    assert report == {"changed": [], "drifted": [], "missing": []}


def test_status_detects_pure_data_edit(tmp_path: Path) -> None:
    root = tmp_path / "gd"
    _build_project(root)
    result = runner.invoke(app, ["export", "--all", "--root", str(root)])
    assert result.exit_code == 0, result.output

    # 只改 Excel 数据（schema 未变）→ 应报数据变更，而非模板漂移
    _set_price(root, 999.0)
    report = canonical_status(root)
    assert "Item" in report["changed"]
    assert report["drifted"] == []


def test_status_detects_schema_drift(tmp_path: Path) -> None:
    root = tmp_path / "gd"
    _build_project(root)
    result = runner.invoke(app, ["export", "--all", "--root", str(root)])
    assert result.exit_code == 0, result.output

    # 改 schema（新增字段）→ 模板漂移；数据未变 → 不归为数据变更
    schema_path = root / "config" / "schemas" / "Item.yaml"
    schema = yaml.safe_load(schema_path.read_text(encoding="utf-8"))
    schema["fields"].append({"name": "Weight", "type": "float"})
    schema_path.write_text(yaml.safe_dump(schema, allow_unicode=True), encoding="utf-8")
    report = canonical_status(root)
    assert "Item" in report["drifted"]
    assert "Item" not in report["changed"]


def test_status_reports_untouched_before_any_export(tmp_path: Path) -> None:
    root = tmp_path / "gd"
    _build_project(root)
    # 未导出：无 canonical-cache/1 状态 → 全部待导出
    report = canonical_status(root)
    assert "Item" in report["changed"]
