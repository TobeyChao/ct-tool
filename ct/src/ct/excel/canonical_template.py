"""Generate Excel template workbooks from the canonical v4 ``Layout``.

The header is a tree rendered from each column's stable canonical path:

- row 1: the top-level field name + its type annotation, merged horizontally
  across the field's column range (primary key cells use the gold fill);
- rows 2..header_rows-1: intermediate segments (record / expanded group
  levels) merged across sibling columns;
- last header row: leaf field comments.

Every column's stable path and annotation is driven by ``ct.excel.layout``,
so headers always match the Web type expressions.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.packaging.custom import DateTimeProperty, IntProperty, StringProperty
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ct.excel.layout import Column, Layout
from ct.schema.resources import EnumResource

_NAME_RUN_FONT = InlineFont(rFont="Segoe UI", b=True, sz=12, color="FFFFFF")
_TYPE_RUN_FONT = InlineFont(rFont="Consolas", i=True, sz=9, color="D8F3DC")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NORMAL_FILL = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
_GROUP_FILL = PatternFill(start_color="40916C", end_color="40916C", fill_type="solid")
_PRIMARY_FILL = PatternFill(start_color="C9A227", end_color="C9A227", fill_type="solid")
_COMMENT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_COMMENT_FONT = Font(name="微软雅黑", italic=True, size=9, color="888888")
_ZEBRA_FILL = PatternFill(start_color="EDF7EE", end_color="EDF7EE", fill_type="solid")
_WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_COL_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)
_NAME_ROW_HEIGHT = 36

_META_TOOL_VERSION = "ct_tool_version"
_META_TABLE_NAME = "ct_table_name"
_META_HEADER_ROWS = "ct_header_rows"
_META_SCHEMA_HASH = "ct_schema_hash"
_META_GENERATED_AT = "ct_generated_at"


def _richtext(name: str, annotation: str) -> CellRichText:
    return CellRichText(
        [
            TextBlock(_NAME_RUN_FONT, f"{name}\n"),
            TextBlock(_TYPE_RUN_FONT, annotation),
        ]
    )


def _top_segment(table_id: str, stable_path: str) -> str:
    """Top-level field name (no group marker) for a column's stable path."""
    tail = stable_path[len(table_id) + 1:]
    head = tail.partition("/")[0]
    return head.partition("[")[0]


def _segments(table_id: str, stable_path: str) -> list[str]:
    """Path segments with each ``[g]`` group marker expanded to its own level."""
    tail = stable_path[len(table_id) + 1:]
    segments: list[str] = []
    for chunk in tail.split("/"):
        if "[" in chunk:
            name, _, rest = chunk.partition("[")
            group = rest.split("]", 1)[0]
            segments.append(name)
            segments.append(group)
        else:
            segments.append(chunk)
    return segments


def generate_canonical_template(
    layout: Layout,
    out_path: Path,
    *,
    enums: dict[str, EnumResource],
    primary: str = "",
) -> Path:
    """Write a v4 template workbook for *layout* and return its path."""
    wb = Workbook()
    ws = wb.active
    ws.title = layout.table_id.partition(":")[2]
    table_name = layout.table_id.partition(":")[2]
    group_rows = layout.header_rows - 1
    comment_row = layout.header_rows

    _write_header_rows(ws, layout, table_name, group_rows, comment_row, primary=primary)

    total_cols = layout.column_count
    for column in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(column)].width = 16
    for row in range(1, group_rows + 1):
        ws.row_dimensions[row].height = _NAME_ROW_HEIGHT
    ws.freeze_panes = ws.cell(row=layout.header_rows + 1, column=1)

    _add_data_validations(ws, layout, enums, data_start=layout.header_rows + 1)
    _add_zebra(ws, layout, data_start=layout.header_rows + 1)

    _write_metadata(wb, layout, table_name)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def _column_ranges(
    layout: Layout,
    table_name: str,
) -> dict[str, tuple[int, int]]:
    ranges: dict[str, list[int]] = {}
    for column in layout.columns:
        top = _top_segment(layout.table_id, column.stable_path)
        ranges.setdefault(top, []).append(column.index)
    return {
        top: (min(indexes), max(indexes))
        for top, indexes in ranges.items()
    }


def _write_header_rows(
    ws,
    layout: Layout,
    table_name: str,
    group_rows: int,
    comment_row: int,
    primary: str = "",
) -> None:
    top_ranges = _column_ranges(layout, table_name)

    # row 1: top-level fields merged across their column range
    top_annotations: dict[str, str] = {}
    for column in layout.columns:
        top = _top_segment(layout.table_id, column.stable_path)
        top_annotations.setdefault(top, column.field_annotation or column.annotation)
    for top, (start, end) in sorted(top_ranges.items()):
        if start < end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start)
        cell.value = _richtext(top, top_annotations[top])
        cell.alignment = _CENTER
        cell.fill = _PRIMARY_FILL if top == primary else _NORMAL_FILL
        cell.border = _THIN
        for column in range(start, end + 1):
            ws.cell(row=1, column=column).border = _THIN

    # rows 2..group_rows: intermediate segments merged across siblings
    for row in range(2, group_rows + 1):
        depth = row
        grouped: dict[tuple[str, ...], list[int]] = {}
        first_by_key: dict[tuple[str, ...], str] = {}
        for column in layout.columns:
            segments = _segments(layout.table_id, column.stable_path)
            if len(segments) < depth:
                continue  # column has already ended above this row
            key = tuple(segments[: depth - 1])
            grouped.setdefault(key, []).append(column.index)
            first_by_key.setdefault(key, segments[depth - 1])
        for key, indexes in sorted(grouped.items()):
            segment = first_by_key[key]
            start_index, end_index = min(indexes), max(indexes)
            if start_index < end_index:
                ws.merge_cells(
                    start_row=row, start_column=start_index,
                    end_row=row, end_column=end_index,
                )
            cell = ws.cell(row=row, column=start_index)
            leaf_annotation = next(
                (
                    column.annotation
                    for column in layout.columns
                    if column.depth == depth and column.index == start_index
                ),
                "",
            )
            cell.value = _richtext(segment, leaf_annotation)
            cell.alignment = _CENTER
            cell.fill = _GROUP_FILL
            cell.border = _THIN
            for column in range(start_index, end_index + 1):
                ws.cell(row=row, column=column).border = _THIN

    # comment row: leaf comments
    for column in layout.columns:
        cell = ws.cell(row=comment_row, column=column.index)
        cell.value = column.comment or ""
        cell.font = _COMMENT_FONT
        cell.alignment = _CENTER
        cell.fill = _COMMENT_FILL
        cell.border = _THIN


def _add_data_validations(ws, layout: Layout, enums: dict[str, EnumResource], data_start: int) -> None:
    for column in layout.columns:
        enum = enums.get(column.type_text)
        if enum is None:
            continue
        formula = '"' + ",".join(enum.values) + '"'
        if len(formula) > 255:
            continue
        letter = get_column_letter(column.index)
        dv = DataValidation(
            type="list",
            formula1=formula,
            showDropDown=False,
            allow_blank=True,
        )
        dv.sqref = f"{letter}{data_start}:{letter}1000"
        ws.add_data_validation(dv)


def _add_zebra(ws, layout: Layout, data_start: int) -> None:
    last = get_column_letter(layout.column_count)
    data_range = f"A{data_start}:{last}1000"
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=_ZEBRA_FILL, border=_COL_BORDER),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=1"], fill=_WHITE_FILL, border=_COL_BORDER),
    )


def _write_metadata(wb: Workbook, layout: Layout, table_name: str) -> None:
    from datetime import datetime, timezone

    props = wb.custom_doc_props
    props.append(StringProperty(name=_META_TOOL_VERSION, value="ct-v4"))
    props.append(StringProperty(name=_META_TABLE_NAME, value=table_name))
    props.append(IntProperty(name=_META_HEADER_ROWS, value=layout.header_rows))
    props.append(StringProperty(name=_META_SCHEMA_HASH, value=layout.schema_hash))
    props.append(
        DateTimeProperty(
            name=_META_GENERATED_AT,
            value=datetime.now(timezone.utc),
        )
    )
