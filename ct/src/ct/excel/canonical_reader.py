"""Read Excel data rows against a canonical v4 ``Layout``.

Values are rebuilt into their canonical shape from each column's stable path:

- a named Record field reassembles its leaf columns into a nested dict;
- an expanded ``vector<Record>`` reads each group's leaf columns in order and
  drops fully-empty trailing groups;
- a single-cell vector splits the cell by its separator.

Parse errors carry the Excel row, column and canonical field path.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ct.excel.layout import Column, Layout
from ct.schema.resources import TableResource
from ct.schema.type_expression import NamedType, ScalarType, VectorType
from ct.validate.errors import IssueCode, ValidationIssue

_BOOL_TRUE = frozenset({"true", "1", "yes", "TRUE", "True", "YES", "Yes", "✓"})
_BOOL_FALSE = frozenset({"false", "0", "no", "FALSE", "False", "NO", "No", "✗"})


def _coerce_scalar(type_text: str, raw: Any) -> tuple[Any, bool]:
    """Coerce one leaf cell; returns (value, ok). None passes through for scalars."""
    if raw is None:
        if type_text == "string":
            return "", True
        return None, True
    if type_text in ("int32", "int64"):
        try:
            return int(float(raw)) if isinstance(raw, float) else int(raw), True
        except (TypeError, ValueError):
            return raw, False
    if type_text in ("float", "double"):
        try:
            return float(raw), True
        except (TypeError, ValueError):
            return raw, False
    if type_text == "bool":
        if isinstance(raw, bool):
            return raw, True
        text = str(raw).strip()
        if text in _BOOL_TRUE:
            return True, True
        if text in _BOOL_FALSE:
            return False, True
        return raw, False
    # enum / named leaves are string identifiers
    return str(raw), True


@dataclass(frozen=True)
class CanonicalParsedRows:
    rows: list[dict[str, Any]]
    excel_rows: list[int]
    issues: list[ValidationIssue] = field(default_factory=list)


def _leaf_path_after_table(table_id: str, stable_path: str) -> list[str]:
    """Segments after the table id, with ``[g]`` markers kept as segments."""
    tail = stable_path[len(table_id) + 1:]
    segments: list[str] = []
    for chunk in tail.split("/"):
        if "[" in chunk:
            name, _, rest = chunk.partition("[")
            segments.append(name)
            segments.append(rest.split("]", 1)[0])
        else:
            segments.append(chunk)
    return segments


class _RowReader:
    def __init__(
        self,
        layout: Layout,
        table: TableResource,
        *,
        records: dict[str, object],
        excel_row: int,
        row_index: int,
    ) -> None:
        self.layout = layout
        self.table = table
        self.records = records
        self.excel_row = excel_row
        self.row_index = row_index
        self.issues: list[ValidationIssue] = []
        self.value_by_path: dict[str, Any] = {}
        self.separators = {
            f"{self.table.resource_id}/{field.name}": field.separator or ","
            for field in self.table.fields
        }

    def read(self, cells: tuple[Any, ...]) -> dict[str, Any] | None:
        if all(
            c is None or (isinstance(c, str) and not c.strip()) for c in cells
        ):
            return None
        for column in self.layout.columns:
            raw = cells[column.index - 1] if column.index - 1 < len(cells) else None
            self.value_by_path[column.stable_path] = raw
        result: dict[str, Any] = {}
        for field in self.table.fields:
            field_value = self._read_field(field, cells)
            if field_value is not None:
                result[field.name] = field_value
        return result

    def _coerce(self, column: Column, raw: Any) -> Any:
        value, ok = _coerce_scalar(column.type_text, raw)
        if not ok:
            self.issues.append(
                ValidationIssue(
                    table=self.table.table,
                    code=IssueCode.TYPE,
                    message=f"期望 {column.type_text} 类型",
                    row_index=self.row_index,
                    excel_row=self.excel_row,
                    column=column.index - 1,
                    field=column.stable_path,
                    value=raw,
                )
            )
        return value

    def _is_record(self, named: NamedType) -> bool:
        return named.name in self.records

    def _read_field(self, field, cells: tuple[Any, ...]) -> Any:
        type_expr = field.type_expr
        owner = self.table.resource_id
        top = f"{owner}/{field.name}"

        if isinstance(type_expr, VectorType) and isinstance(
            type_expr.element, NamedType
        ) and self._is_record(type_expr.element) and (field.excel_columns or 0) > 0:
            groups: list[dict[str, Any]] = []
            group_count = field.excel_columns or 0
            for group in range(1, group_count + 1):
                element = self._read_record_group(type_expr.element, top, group)
                if element is None:
                    break  # groups are contiguous; first empty group ends the list
                groups.append(element)
            return groups
        if isinstance(type_expr, VectorType):
            column = next(
                column
                for column in self.layout.columns
                if column.stable_path == top
            )
            raw = self.value_by_path.get(top)
            return self._split_vector(type_expr, column, raw)
        if isinstance(type_expr, NamedType):
            if self._is_record(type_expr):
                return self._read_record(type_expr, top)
            column = next(
                column for column in self.layout.columns if column.stable_path == top
            )
            return self._coerce(column, self.value_by_path.get(top))
        column = next(
            column for column in self.layout.columns if column.stable_path == top
        )
        return self._coerce(column, self.value_by_path.get(top))

    def _record_leaf_columns(self, top: str, group: int | None) -> list[Column]:
        return [
            column
            for column in self.layout.columns
            if column.stable_path.startswith(top)
            and (group is None or column.group_index == group)
        ]

    def _read_record(self, named: NamedType, top: str) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for column in self._record_leaf_columns(top, group=None):
            tail = _leaf_path_after_table(self.layout.table_id, column.stable_path)
            leaf = tail[-1]
            result[leaf] = self._coerce(column, self.value_by_path.get(column.stable_path))
        return result

    def _read_record_group(self, named: NamedType, top: str, group: int) -> dict[str, Any] | None:
        columns = [
            column
            for column in self._record_leaf_columns(top, group=group)
            if column.group_index == group
        ]
        if not columns:
            return None
        if all(
            self.value_by_path.get(column.stable_path) is None
            for column in columns
        ):
            return None  # fully-empty group
        result: dict[str, Any] = {}
        for column in columns:
            tail = _leaf_path_after_table(self.layout.table_id, column.stable_path)
            leaf = tail[-1]
            result[leaf] = self._coerce(column, self.value_by_path.get(column.stable_path))
        return result

    def _split_vector(self, vector: VectorType, column: Column, raw: Any) -> list[Any]:
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            return []
        separator = self.separators.get(column.stable_path, ",")
        element_text = self._vector_element_text(vector)
        elements: list[Any] = []
        for token in str(raw).split(separator):
            token = token.strip()
            if not token:
                continue
            value, ok = _coerce_scalar(element_text, token)
            if not ok:
                self.issues.append(
                    ValidationIssue(
                        table=self.table.table,
                        code=IssueCode.TYPE,
                        message=f"第{len(elements) + 1}个元素期望 {element_text} 类型",
                        row_index=self.row_index,
                        excel_row=self.excel_row,
                        column=column.index - 1,
                        field=column.stable_path,
                        value=token,
                    )
                )
            elements.append(value)
        return elements

    def _vector_element_text(self, vector: VectorType) -> str:
        from ct.schema.type_expression import serialize_type_expression

        return serialize_type_expression(vector.element)


def read_canonical_excel(
    excel_path: Path,
    layout: Layout,
    table: TableResource,
    *,
    records: dict[str, object] | None = None,
) -> CanonicalParsedRows:
    """Read Excel rows against a canonical layout; returns canonical values."""
    wb = load_workbook(str(excel_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return CanonicalParsedRows([], [])
        rows: list[dict[str, Any]] = []
        excel_rows: list[int] = []
        issues: list[ValidationIssue] = []
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index <= layout.header_rows:
                continue
            reader = _RowReader(
                layout,
                table,
                records=records,
                excel_row=row_index,
                row_index=len(rows) + 1,
            )
            parsed = reader.read(tuple(row))
            if parsed is not None:
                rows.append(parsed)
                excel_rows.append(row_index)
                issues.extend(reader.issues)
        return CanonicalParsedRows(rows, excel_rows, issues)
    finally:
        wb.close()
