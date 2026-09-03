"""Excel data change planning: stable-path migration + risk scans.

The Change Plan compares the old layout (from the persisted manifest or the
old layout model) with a candidate layout, maps columns by stable logical
path (with explicit renames taking precedence), and scans actual non-empty
cells for deleted columns, ``excel_columns`` shrink, enum value removal and
type conversion. Anything that cannot be proven lossless blocks Apply.
Untracked workbooks (missing/corrupt manifest) never get a silent writeback.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ct.excel.canonical_reader import _coerce_scalar
from ct.excel.layout import Column, Layout
from ct.excel.layout_manifest import LayoutManifest


@dataclass(frozen=True)
class PlanIssue:
    kind: str  # "blocker" | "untracked" | "warning"
    message: str
    rows: tuple[int, ...] = ()
    columns: tuple[int, ...] = ()
    field_path: str = ""
    samples: tuple[Any, ...] = ()

    def render(self) -> str:
        location = ""
        if self.rows:
            location = f"（Excel 行 {self.rows} · 列 {tuple(c + 1 for c in self.columns)}）"
        samples = ""
        if self.samples:
            samples = f"，样例 {self.samples}"
        return f"[{self.kind}] {self.message}{location}{samples}"


@dataclass(frozen=True)
class ColumnMigration:
    old_index: int  # 1-based old column
    new_index: int | None  # 1-based new column; None = deleted
    old_path: str
    new_path: str | None = None


@dataclass(frozen=True)
class ExcelPlan:
    migrations: tuple[ColumnMigration, ...]
    issues: tuple[PlanIssue, ...]
    untracked: bool = False

    @property
    def blocked(self) -> bool:
        return any(issue.kind == "blocker" for issue in self.issues) or self.untracked


def _read_data_rows(path: Path, header_rows: int) -> list[tuple[int, tuple[Any, ...]]]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        return [
            (index, tuple(row))
            for index, row in enumerate(ws.iter_rows(values_only=True), start=1)
            if index > header_rows
            and any(
                cell is not None
                and not (isinstance(cell, str) and not cell.strip())
                for cell in row
            )
        ]
    finally:
        wb.close()


def _logical_map(columns: tuple[Column, ...]) -> dict[str, Column]:
    return {column.logical_path: column for column in columns}


def _coerce_ok(type_text: str, raw: Any) -> bool:
    _, ok = _coerce_scalar(type_text, raw)
    return ok


def plan_excel_migration(
    old_layout: Layout,
    new_layout: Layout,
    excel_path: Path,
    *,
    rename_map: dict[str, str] | None = None,
    new_enums: dict[str, tuple[str, ...]] | None = None,
    old_enums: dict[str, tuple[str, ...]] | None = None,
    manifest: LayoutManifest | None = None,
) -> ExcelPlan:
    """Plan how to move *excel_path* data from *old_layout* to *new_layout*."""
    rename_map = rename_map or {}
    untracked = manifest is None
    old_manifest = manifest or LayoutManifest.from_layout(old_layout)

    old_logical = _logical_map(old_layout.columns)
    new_logical = _logical_map(new_layout.columns)

    data_rows = _read_data_rows(excel_path, old_manifest.header_rows)
    issues: list[PlanIssue] = []

    migrations: list[ColumnMigration] = []
    new_used: set[int] = set()
    for old_column in old_layout.columns:
        mapped = rename_map.get(old_column.logical_path, old_column.logical_path)
        target = new_logical.get(mapped)
        if target is None or target.index in new_used:
            # deleted, or collapsed by a previous old column (excel_columns shrink)
            migrations.append(
                ColumnMigration(old_column.index, None, old_column.stable_path)
            )
            _scan_deleted(old_column, data_rows, issues)
            continue
        new_used.add(target.index)
        migrations.append(
            ColumnMigration(
                old_column.index, target.index, old_column.stable_path, target.stable_path
            )
        )
        _scan_type_change(old_column, target, data_rows, issues)

    _scan_enum_removal(
        old_layout.columns,
        new_layout.columns,
        data_rows,
        issues,
        old_enums=old_enums,
        new_enums=new_enums,
    )

    if untracked:
        issues.append(
            PlanIssue(
                kind="untracked",
                message="Excel 缺少可信路径清单（template_layouts manifest），"
                "列映射需人工核对，不允许静默写回",
            )
        )

    return ExcelPlan(
        migrations=tuple(migrations),
        issues=tuple(issues),
        untracked=untracked,
    )


def _scan_deleted(
    column: Column,
    rows: list[tuple[int, tuple[Any, ...]]],
    issues: list[PlanIssue],
) -> None:
    populated = [
        (excel_row, row[column.index - 1])
        for excel_row, row in rows
        if column.index - 1 < len(row) and row[column.index - 1] is not None
    ]
    if populated:
        issues.append(
            PlanIssue(
                kind="blocker",
                message=f"字段 {column.stable_path} 被删除但存在非空数据",
                rows=tuple(excel_row for excel_row, _ in populated[:5]),
                columns=(column.index - 1,),
                field_path=column.stable_path,
                samples=tuple(str(value) for _, value in populated[:3]),
            )
        )


def _scan_type_change(
    old_column: Column,
    new_column: Column,
    rows: list[tuple[int, tuple[Any, ...]]],
    issues: list[PlanIssue],
) -> None:
    if old_column.type_text == new_column.type_text:
        return
    failed = [
        (excel_row, row[old_column.index - 1])
        for excel_row, row in rows
        if old_column.index - 1 < len(row)
        and row[old_column.index - 1] is not None
        and not _coerce_ok(new_column.type_text, row[old_column.index - 1])
    ]
    if failed:
        issues.append(
            PlanIssue(
                kind="blocker",
                message=(
                    f"字段 {old_column.stable_path} 类型 {old_column.type_text} → "
                    f"{new_column.type_text}，存在不可转换值"
                ),
                rows=tuple(excel_row for excel_row, _ in failed[:5]),
                columns=(old_column.index - 1,),
                field_path=old_column.stable_path,
                samples=tuple(str(value) for _, value in failed[:3]),
            )
        )


def _scan_enum_removal(
    old_columns: tuple[Column, ...],
    new_columns: tuple[Column, ...],
    rows: list[tuple[int, tuple[Any, ...]]],
    issues: list[PlanIssue],
    *,
    old_enums: dict[str, tuple[str, ...]] | None,
    new_enums: dict[str, tuple[str, ...]] | None,
) -> None:
    if not old_enums or not new_enums:
        return
    new_logical = _logical_map(new_columns)
    for column in old_columns:
        old_values = old_enums.get(column.type_text)
        if old_values is None:
            continue
        new_values = new_enums.get(column.type_text, ())
        removed = [value for value in old_values if value not in new_values]
        if not removed:
            continue
        if column.logical_path not in new_logical:
            continue  # column deleted; already reported
        populated = [
            (excel_row, row[column.index - 1])
            for excel_row, row in rows
            if column.index - 1 < len(row)
            and row[column.index - 1] in removed
        ]
        if populated:
            issues.append(
                PlanIssue(
                    kind="blocker",
                    message=(
                        f"Enum {column.type_text} 移除了值 "
                        f"{', '.join(removed)}，字段 {column.stable_path} 存在旧值"
                    ),
                    rows=tuple(excel_row for excel_row, _ in populated[:5]),
                    columns=(column.index - 1,),
                    field_path=column.stable_path,
                    samples=tuple(str(value) for _, value in populated[:3]),
                )
            )
