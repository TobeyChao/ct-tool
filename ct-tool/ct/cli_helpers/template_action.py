"""Decision logic for the ``ct gen-template`` command.

Centralizes the matrix of (file state × flags) → action so that the CLI
loop only needs to dispatch the resulting action verb.
"""

from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from pathlib import Path

from ct.excel.template import read_template_metadata
from ct.schema.hashing import compute_schema_hash
from ct.schema.models import TableSchema


class Action(str, Enum):
    CREATE_NEW = "create_new"          # file doesn't exist
    SKIP = "skip"                      # hash matches, no flag
    REBUILD = "rebuild"                # full overwrite
    UPDATE_PRESERVE = "update_preserve"  # rebuild header, keep data
    REFUSE = "refuse"                  # block with explanation


@dataclass(frozen=True)
class Decision:
    action: Action
    message: str  # human-readable explanation; empty for silent paths


def _has_data_rows(path: Path, header_rows: int) -> bool:
    """Return True if the workbook has at least one non-empty row past the header."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        # Treat unreadable file as "has data" — safer to refuse than to silently overwrite.
        return True
    try:
        ws = wb.active
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx <= header_rows:
                continue
            if any(c is not None for c in row):
                return True
    finally:
        wb.close()
    return False


def decide_template_action(
    schema: TableSchema,
    path: Path,
    *,
    force: bool,
    update_header: bool,
) -> Decision:
    """Return what to do with *path* given the schema and the user's flags.

    See specs/cli-interface decision matrix. ``--force`` and
    ``--update-header`` are mutually compatible flags; if both are given,
    ``--update-header`` wins (data-preserving path is the safer choice).
    """
    table = schema.table

    # ---- File doesn't exist → simple create.
    if not path.exists():
        return Decision(Action.CREATE_NEW, f"[new] {path.name}")

    meta = read_template_metadata(path)

    # ---- table_name mismatch is the one path with no escape hatch.
    if meta is not None and meta.table_name != table:
        return Decision(
            Action.REFUSE,
            f"[refuse] {path.name} 元数据归属为 '{meta.table_name}'，与当前 schema "
            f"'{table}' 不一致。请手动确认归属后重试（改名或删除文件）。"
            f"任何 flag（--force / --update-header）都不会绕过此检查。",
        )

    # ---- Legacy file: no metadata.
    if meta is None:
        if update_header:
            return Decision(
                Action.UPDATE_PRESERVE,
                f"[update] {table}: 文件无元数据（legacy），用当前 schema 的 "
                f"header_rows={schema.header_rows} 推断旧表头。请检查首尾行是否被误跳/误带。",
            )
        if force:
            return Decision(
                Action.REBUILD,
                f"[force] {table}: 全量覆盖 legacy 文件（数据将丢失）。",
            )
        return Decision(
            Action.REFUSE,
            f"[refuse] {table}: 文件无元数据（可能由旧版工具或手工创建）。"
            f"使用 --update-header 用当前 schema 推断保留数据，或 --force 强制覆盖。",
        )

    # ---- Has metadata: compare hashes.
    current_hash = compute_schema_hash(schema)
    if meta.schema_hash == current_hash:
        if update_header:
            # Hash matched but user explicitly wants to rebuild while preserving data.
            return Decision(
                Action.UPDATE_PRESERVE,
                f"[update] {table}: schema 未变化但已按 --update-header 重建。",
            )
        if force:
            return Decision(
                Action.REBUILD,
                f"[force] {table}: schema 未变化，但已按 --force 全量重建（数据将丢失）。",
            )
        return Decision(
            Action.SKIP,
            f"[skip] {table}: schema 未变化，模板无需重建（如需强制重建请加 --force）。",
        )

    # ---- Hash differs: schema has changed.
    if update_header:
        return Decision(
            Action.UPDATE_PRESERVE,
            f"[update] {table}: schema 已修改，重建表头并保留数据。",
        )

    has_data = _has_data_rows(path, meta.header_rows)
    if not has_data:
        # Safe to rebuild even without flags — nothing to lose.
        return Decision(
            Action.REBUILD,
            f"[rebuild] {table}: schema 已修改，文件无数据，已直接重建。",
        )

    if force:
        return Decision(
            Action.REBUILD,
            f"[force] {table}: schema 已修改，已按 --force 全量覆盖（数据将丢失）。",
        )

    return Decision(
        Action.REFUSE,
        f"[refuse] {table}: schema 已修改且文件含数据。"
        f"使用 --update-header 保留数据重建表头，或 --force 强制覆盖。",
    )