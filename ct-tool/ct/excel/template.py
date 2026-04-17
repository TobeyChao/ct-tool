"""Generate Excel template workbooks with structured, multi-row headers
derived from a ``TableSchema``.

Header layout (total rows = ``schema.header_rows`` = max_nesting_depth + 2):

1. The first ``max_nesting_depth`` rows are **group / name rows**.
   - Non-struct fields write their name in the first group row and merge
     vertically downward through all remaining group rows.
   - Struct fields write the struct name in their group row, merged
     horizontally across the span of their sub-fields, then recurse into
     the next group row for sub-field names.
2. The **type row** (second-to-last) shows type annotations such as
   ``enum[v1,v2]``, ``int32[ref:table]``, ``string[i18n]``,
   ``array<int32>``.
3. The **comment row** (last) shows ``field.comment``.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from importlib import metadata as importlib_metadata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.packaging.custom import (
    DateTimeProperty,
    IntProperty,
    StringProperty,
)
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ct.schema.hashing import compute_schema_hash
from ct.schema.models import FieldDef, TableSchema

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

# Name-row fonts: white bold on dark backgrounds
_HEADER_FONT = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
# Annotation-row fonts
_TYPE_FONT = Font(name="Consolas", italic=True, size=10, color="1B4332")
_COMMENT_FONT = Font(name="微软雅黑", italic=True, size=9, color="888888")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Header row fills — deep green hierarchy (dark → mid → annotation)
_NORMAL_FILL  = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")  # deep forest green
_STRUCT_FILL  = PatternFill(start_color="40916C", end_color="40916C", fill_type="solid")  # mid green
_PRIMARY_FILL = PatternFill(start_color="C9A227", end_color="C9A227", fill_type="solid")  # warm gold (primary key)
_TYPE_FILL    = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")  # pale green
_COMMENT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # light gray

# Data-row zebra fills
_ZEBRA_FILL = PatternFill(start_color="EDF7EE", end_color="EDF7EE", fill_type="solid")  # pale green (even rows)
_WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # white (odd rows)
_COL_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)


# ---------------------------------------------------------------------------
# Column span (same logic as reader, duplicated to keep modules independent)
# ---------------------------------------------------------------------------

def _column_span(field: FieldDef) -> int:
    if field.type == "struct" and field.fields:
        return sum(_column_span(sf) for sf in field.fields)
    return 1


# ---------------------------------------------------------------------------
# Collect all leaf fields with their 1-based column positions
# ---------------------------------------------------------------------------

def _collect_leaf_fields(fields: list[FieldDef], start_col: int) -> list[tuple[FieldDef, int]]:
    """Return (leaf_field, col_index) pairs for all non-struct leaf fields."""
    result: list[tuple[FieldDef, int]] = []
    col = start_col
    for field in fields:
        if field.type == "struct" and field.fields:
            result.extend(_collect_leaf_fields(field.fields, col))
            col += _column_span(field)
        else:
            result.append((field, col))
            col += 1
    return result


# ---------------------------------------------------------------------------
# Type annotation string
# ---------------------------------------------------------------------------

def _type_annotation(field: FieldDef) -> str:
    """Build a human-readable type annotation for the type row."""
    base = field.type

    if field.type == "enum":
        vals = ",".join(field.values or [])
        return f"enum[{vals}]"

    if field.type == "array":
        element = field.element or "?"
        if element == "enum":
            vals = ",".join(field.element_values or [])
            return f"array<enum[{vals}]>"
        return f"array<{element}>"

    # Basic types with optional qualifiers
    parts: list[str] = []
    if field.ref:
        parts.append(f"ref:{field.ref}")
    if field.i18n:
        parts.append("i18n")
    if parts:
        return f"{base}[{','.join(parts)}]"
    return base


# ---------------------------------------------------------------------------
# Recursive header writer
# ---------------------------------------------------------------------------

def _write_field_headers(
    ws,
    fields: list[FieldDef],
    start_col: int,
    group_row_start: int,
    group_row_end: int,
    type_row: int,
    comment_row: int,
    primary_key: str = "",
) -> int:
    """Write header cells for *fields* starting at *start_col*.

    Returns the next available column index (1-based).

    Parameters
    ----------
    ws : Worksheet
    fields : list of field definitions at this nesting level
    start_col : 1-based column index to start writing
    group_row_start : 1-based row for this nesting level's names
    group_row_end : 1-based row of the last group row (before the type row)
    type_row : 1-based row index for type annotations
    comment_row : 1-based row index for comments
    """
    col = start_col
    for field in fields:
        span = _column_span(field)

        if field.type == "struct" and field.fields:
            # ---- struct: horizontal merge in current group row ----
            end_col = col + span - 1
            if span > 1:
                ws.merge_cells(
                    start_row=group_row_start, start_column=col,
                    end_row=group_row_start, end_column=end_col,
                )
            cell = ws.cell(row=group_row_start, column=col, value=field.name)
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.fill = _STRUCT_FILL
            cell.border = _THIN_BORDER

            for c in range(col, end_col + 1):
                ws.cell(row=group_row_start, column=c).border = _THIN_BORDER

            # Recurse into sub-fields one row deeper
            _write_field_headers(
                ws, field.fields,
                start_col=col,
                group_row_start=group_row_start + 1,
                group_row_end=group_row_end,
                type_row=type_row,
                comment_row=comment_row,
                primary_key=primary_key,
            )
            col = end_col + 1

        else:
            # ---- leaf field: vertical merge across remaining group rows ----
            is_primary = field.name == primary_key
            name_fill = _PRIMARY_FILL if is_primary else _NORMAL_FILL

            if group_row_start < group_row_end:
                ws.merge_cells(
                    start_row=group_row_start, start_column=col,
                    end_row=group_row_end, end_column=col,
                )
            cell = ws.cell(row=group_row_start, column=col, value=field.name)
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.fill = name_fill
            cell.border = _THIN_BORDER

            for r in range(group_row_start, group_row_end + 1):
                ws.cell(row=r, column=col).border = _THIN_BORDER

            # Type row
            type_cell = ws.cell(row=type_row, column=col,
                                value=_type_annotation(field))
            type_cell.font = _TYPE_FONT
            type_cell.alignment = _CENTER
            type_cell.fill = _TYPE_FILL
            type_cell.border = _THIN_BORDER

            # Comment row
            comment_cell = ws.cell(row=comment_row, column=col,
                                   value=field.comment or "")
            comment_cell.font = _COMMENT_FONT
            comment_cell.alignment = _CENTER
            comment_cell.fill = _COMMENT_FILL
            comment_cell.border = _THIN_BORDER

            col += 1

    return col


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

# Metadata field names stored in Excel Custom Document Properties.
_META_TOOL_VERSION = "ct_tool_version"
_META_TABLE_NAME = "ct_table_name"
_META_HEADER_ROWS = "ct_header_rows"
_META_SCHEMA_HASH = "ct_schema_hash"
_META_GENERATED_AT = "ct_generated_at"

_META_REQUIRED = (
    _META_TOOL_VERSION,
    _META_TABLE_NAME,
    _META_HEADER_ROWS,
    _META_SCHEMA_HASH,
    _META_GENERATED_AT,
)


@dataclass(frozen=True)
class TemplateMetadata:
    """Self-describing metadata embedded in a generated Excel template."""

    tool_version: str
    table_name: str
    header_rows: int
    schema_hash: str
    generated_at: str


def _tool_version() -> str:
    try:
        return importlib_metadata.version("ct-tool")
    except importlib_metadata.PackageNotFoundError:
        return "unknown"


def _write_metadata(wb: Workbook, schema: TableSchema) -> None:
    """Write the five ct_* properties to the workbook's custom doc props."""
    props = wb.custom_doc_props
    # Drop any pre-existing ct_* entries so we never end up with duplicates.
    for name in _META_REQUIRED:
        try:
            del props[name]
        except (KeyError, AttributeError):
            pass

    now = datetime.now(timezone.utc)
    props.append(StringProperty(name=_META_TOOL_VERSION, value=_tool_version()))
    props.append(StringProperty(name=_META_TABLE_NAME, value=schema.table))
    props.append(IntProperty(name=_META_HEADER_ROWS, value=schema.header_rows))
    props.append(StringProperty(name=_META_SCHEMA_HASH, value=compute_schema_hash(schema)))
    props.append(DateTimeProperty(name=_META_GENERATED_AT, value=now))


def read_template_metadata(path: Path) -> TemplateMetadata | None:
    """Read ct_* metadata from an existing Excel file.

    Returns ``None`` for any failure mode: file missing, file corrupted,
    metadata missing, partial metadata, or unexpected types. The caller
    should treat ``None`` as "untracked / legacy" and fall back accordingly.
    """
    if not path.exists():
        return None
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        props = wb.custom_doc_props
        values: dict[str, object] = {}
        for prop in props:
            if prop.name in _META_REQUIRED:
                values[prop.name] = prop.value
        if not all(name in values for name in _META_REQUIRED):
            return None
        try:
            header_rows = int(values[_META_HEADER_ROWS])  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None
        generated_at = values[_META_GENERATED_AT]
        if isinstance(generated_at, datetime):
            generated_at_str = generated_at.isoformat()
        else:
            generated_at_str = str(generated_at)
        return TemplateMetadata(
            tool_version=str(values[_META_TOOL_VERSION]),
            table_name=str(values[_META_TABLE_NAME]),
            header_rows=header_rows,
            schema_hash=str(values[_META_SCHEMA_HASH]),
            generated_at=generated_at_str,
        )
    except Exception:
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def generate_template(schema: TableSchema, output_path: Path) -> None:
    """Create an Excel template workbook with structured headers."""
    total_rows = schema.header_rows  # max_nesting_depth + 2
    group_rows = schema.max_nesting_depth
    type_row = group_rows + 1
    comment_row = group_rows + 2

    wb = Workbook()
    ws = wb.active
    ws.title = schema.table  # type: ignore[union-attr]

    _write_field_headers(
        ws,
        schema.fields,
        start_col=1,
        group_row_start=1,
        group_row_end=group_rows,
        type_row=type_row,
        comment_row=comment_row,
        primary_key=schema.primary,
    )

    total_cols = sum(_column_span(f) for f in schema.fields)
    last_col_letter = get_column_letter(total_cols)

    # Fixed column width
    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16  # type: ignore[union-attr]

    # Freeze panes below the header
    ws.freeze_panes = ws.cell(row=total_rows + 1, column=1)  # type: ignore[union-attr]

    # DataValidation for enum fields
    data_start = total_rows + 1
    for field, col_idx in _collect_leaf_fields(schema.fields, 1):
        if field.type != "enum" or not field.values:
            continue
        formula_str = '"' + ",".join(field.values) + '"'
        if len(formula_str) > 255:
            logger.warning(
                "enum 字段 '%s' 的值列表超过 255 字符，跳过下拉菜单", field.name
            )
            continue
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=formula_str,
            showDropDown=False,
            allow_blank=True,
        )
        dv.sqref = f"{col_letter}{data_start}:{col_letter}1000"
        ws.add_data_validation(dv)  # type: ignore[union-attr]

    # Zebra striping via conditional formatting (fill + column separator borders)
    data_range = f"A{data_start}:{last_col_letter}1000"
    ws.conditional_formatting.add(  # type: ignore[union-attr]
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=_ZEBRA_FILL, border=_COL_BORDER),
    )
    ws.conditional_formatting.add(  # type: ignore[union-attr]
        data_range,
        FormulaRule(formula=["MOD(ROW(),2)=1"], fill=_WHITE_FILL, border=_COL_BORDER),
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_metadata(wb, schema)
    wb.save(str(output_path))
    logger.info("模板已生成: %s (%d 列, %d 行表头)",
                output_path, total_cols, total_rows)


def update_template(schema: TableSchema, output_path: Path) -> int:
    """Rebuild the header rows of an existing template, preserving data rows.

    Reads the file's metadata to determine where the old header ended, captures
    every non-empty row below it, regenerates the template (which writes fresh
    metadata), and appends the captured rows back below the new header.

    Returns the number of data rows that were preserved.

    For legacy files (no metadata) the current schema's ``header_rows`` is used
    as the best guess for where the old header ended — if the old schema's
    nesting depth differed, the caller is expected to warn the user.
    """
    meta = read_template_metadata(output_path)
    old_header_rows = meta.header_rows if meta is not None else schema.header_rows

    # 1. Snapshot data rows from the existing file.
    data_rows: list[tuple] = []
    old_wb = load_workbook(str(output_path), read_only=True, data_only=True)
    try:
        old_ws = old_wb.active
        for idx, row in enumerate(old_ws.iter_rows(values_only=True), start=1):
            if idx <= old_header_rows:
                continue
            if any(cell is not None for cell in row):
                data_rows.append(row)
    finally:
        old_wb.close()

    # 2. Regenerate the template with fresh headers + metadata.
    generate_template(schema, output_path)

    # 3. Append preserved data rows under the new header.
    if data_rows:
        new_wb = load_workbook(str(output_path))
        try:
            new_ws = new_wb.active
            for row in data_rows:
                new_ws.append(list(row))
            new_wb.save(str(output_path))
        finally:
            new_wb.close()

    logger.info(
        "模板已更新: %s (保留 %d 行数据，旧表头 %d 行)",
        output_path, len(data_rows), old_header_rows,
    )
    return len(data_rows)