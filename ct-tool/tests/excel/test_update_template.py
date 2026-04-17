"""Tests for ct.excel.template.update_template."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ct.excel.template import (
    generate_template,
    read_template_metadata,
    update_template,
)
from ct.schema.models import FieldDef, TableSchema


def _schema_v1() -> TableSchema:
    return TableSchema(
        table="item",
        primary="id",
        fields=[
            FieldDef(name="id", type="int32", comment="主键"),
            FieldDef(name="name", type="string", comment="名称"),
            FieldDef(name="price", type="int32", comment="价格"),
        ],
    )


def _schema_v2_with_extra_field() -> TableSchema:
    """Same nesting depth as v1, just one extra column."""
    s = _schema_v1()
    s.fields.append(FieldDef(name="rarity", type="enum", values=["common", "rare"]))
    return s


def _fill_data_rows(path: Path, schema: TableSchema, rows: list[tuple]) -> None:
    """Append data rows below the schema's header to simulate filled template."""
    wb = load_workbook(str(path))
    ws = wb.active
    start = schema.header_rows + 1
    for r_offset, row in enumerate(rows):
        for c_offset, value in enumerate(row):
            ws.cell(row=start + r_offset, column=c_offset + 1, value=value)
    wb.save(str(path))


def _read_data_rows(path: Path, header_rows: int) -> list[tuple]:
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    out: list[tuple] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx <= header_rows:
            continue
        if any(c is not None for c in row):
            out.append(row)
    wb.close()
    return out


def test_update_preserves_data_rows(tmp_path: Path) -> None:
    """Spec scenario: 'Update header preserves data rows'."""
    schema_v1 = _schema_v1()
    out = tmp_path / "item.xlsx"
    generate_template(schema_v1, out)

    seed = [
        (1, "sword", 100),
        (2, "shield", 80),
        (3, "potion", 5),
    ]
    _fill_data_rows(out, schema_v1, seed)

    schema_v2 = _schema_v2_with_extra_field()
    preserved = update_template(schema_v2, out)
    assert preserved == 3

    rows_after = _read_data_rows(out, schema_v2.header_rows)
    assert len(rows_after) == 3
    # Old data is appended verbatim — first 3 columns match seed exactly.
    for original, after in zip(seed, rows_after):
        assert tuple(after[:3]) == original

    # Metadata reflects the new schema.
    meta = read_template_metadata(out)
    assert meta is not None
    assert meta.header_rows == schema_v2.header_rows


def test_update_legacy_file_uses_current_schema_header_rows(tmp_path: Path) -> None:
    """Spec scenario: 'Legacy file uses new schema header_rows'.

    Build a workbook with NO ct_* metadata, fill data below where the new
    schema's header would end, then verify update_template preserves rows.
    """
    from openpyxl import Workbook

    schema = _schema_v1()
    out = tmp_path / "legacy.xlsx"

    # Build a "legacy" workbook with the right number of header rows but no metadata.
    wb = Workbook()
    ws = wb.active
    ws.title = schema.table  # type: ignore[union-attr]
    # Fake header rows (just put strings, no fancy formatting).
    for r in range(1, schema.header_rows + 1):
        ws.cell(row=r, column=1, value=f"hdr-r{r}")
    # Data rows below.
    seed = [(10, "legacy-sword", 200), (11, "legacy-shield", 150)]
    for r_offset, row in enumerate(seed):
        for c_offset, value in enumerate(row):
            ws.cell(row=schema.header_rows + 1 + r_offset, column=c_offset + 1, value=value)
    wb.save(str(out))

    # Sanity check: the legacy file has no metadata.
    assert read_template_metadata(out) is None

    preserved = update_template(schema, out)
    assert preserved == 2

    rows_after = _read_data_rows(out, schema.header_rows)
    assert len(rows_after) == 2
    for original, after in zip(seed, rows_after):
        assert tuple(after[:3]) == original

    # File now carries metadata.
    meta = read_template_metadata(out)
    assert meta is not None
    assert meta.table_name == "item"


def test_update_with_no_data_rows_is_clean_rebuild(tmp_path: Path) -> None:
    schema = _schema_v1()
    out = tmp_path / "empty.xlsx"
    generate_template(schema, out)

    schema_v2 = _schema_v2_with_extra_field()
    preserved = update_template(schema_v2, out)
    assert preserved == 0
    meta = read_template_metadata(out)
    assert meta is not None
    assert meta.header_rows == schema_v2.header_rows